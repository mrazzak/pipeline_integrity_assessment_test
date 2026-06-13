from __future__ import annotations

import json
import mimetypes
import os
import secrets
import hashlib
import hmac
import math
import random
import csv
from datetime import date, datetime, timedelta, timezone
from io import BytesIO, StringIO
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .calculator import calculate_case
from .models import CrossingCase, Pipe, PointLoad, Scan, Soil, Vehicle


ROOT = Path(__file__).resolve().parents[2]
WEB_ROOT = ROOT / "web"
DATA_ROOT = Path(os.environ.get("PIPELINE_ASSESSMENT_DATA_DIR", ROOT / "data")).resolve()
USERS_FILE = DATA_ROOT / "users.json"
METHOD_FILE = DATA_ROOT / "calculation_method.json"
REPORT_ROOT = DATA_ROOT / "reports"
PASSWORD_MAX_AGE_DAYS = 120
PASSWORD_ITERATIONS = 200_000
DEFAULT_CSA_Z662_DESIGN_FACTOR = 0.800
CSA_Z662_LOCATION_FACTORS = {
    "1": 1.000,
    "2": 0.900,
    "3": 0.700,
    "4": 0.550,
}


def today_utc() -> date:
    return datetime.now(timezone.utc).date()


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def iso_date(value: date) -> str:
    return value.isoformat()


def parse_iso_date(value: Any) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    return date.fromisoformat(text)


def password_expiry_from(start: date | None = None) -> str:
    return iso_date((start or today_utc()) + timedelta(days=PASSWORD_MAX_AGE_DAYS))


def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), bytes.fromhex(salt), PASSWORD_ITERATIONS)
    return f"pbkdf2_sha256${PASSWORD_ITERATIONS}${salt}${digest.hex()}"


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        algorithm, iterations, salt, digest = stored_hash.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        candidate = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), bytes.fromhex(salt), int(iterations))
        return hmac.compare_digest(candidate.hex(), digest)
    except (ValueError, TypeError):
        return False


def validate_password_strength(password: str) -> None:
    if len(password) < 8:
        raise ValueError("Password must be at least 8 characters long.")


def validate_email(email: str) -> None:
    if email and ("@" not in email or "." not in email.rsplit("@", 1)[-1]):
        raise ValueError("Enter a valid email address.")


def make_user_record(
    password: str,
    role: str = "user",
    account_expires_at: str = "",
    full_name: str = "",
    email: str = "",
) -> dict[str, Any]:
    validate_password_strength(password)
    changed_at = today_utc()
    return {
        "password_hash": hash_password(password),
        "role": role,
        "full_name": full_name,
        "email": email,
        "account_expires_at": account_expires_at,
        "password_changed_at": iso_date(changed_at),
        "password_expires_at": password_expiry_from(changed_at),
        "last_login_at": "",
        "session_count": 0,
        "total_session_seconds": 0.0,
        "modules_used": [],
        "active_sessions": {},
    }


def normalize_user_record(record: dict[str, Any]) -> tuple[dict[str, Any], bool]:
    changed = False
    normalized = dict(record)
    if "password_hash" not in normalized:
        password = str(normalized.pop("password", ""))
        if password:
            normalized["password_hash"] = hash_password(password)
            changed = True
    if "password" in normalized:
        normalized.pop("password", None)
        changed = True
    if normalized.get("role") not in {"user", "admin"}:
        normalized["role"] = "user"
        changed = True
    if "full_name" not in normalized:
        normalized["full_name"] = ""
        changed = True
    if "email" not in normalized:
        normalized["email"] = ""
        changed = True
    if "account_expires_at" not in normalized:
        normalized["account_expires_at"] = ""
        changed = True
    if "password_changed_at" not in normalized:
        normalized["password_changed_at"] = iso_date(today_utc())
        changed = True
    if "password_expires_at" not in normalized:
        changed_at = parse_iso_date(normalized.get("password_changed_at")) or today_utc()
        normalized["password_expires_at"] = password_expiry_from(changed_at)
        changed = True
    if "last_login_at" not in normalized:
        normalized["last_login_at"] = ""
        changed = True
    if "session_count" not in normalized:
        normalized["session_count"] = 0
        changed = True
    if "total_session_seconds" not in normalized:
        normalized["total_session_seconds"] = 0.0
        changed = True
    if not isinstance(normalized.get("modules_used"), list):
        normalized["modules_used"] = []
        changed = True
    if not isinstance(normalized.get("active_sessions"), dict):
        normalized["active_sessions"] = {}
        changed = True
    return normalized, changed


def account_is_expired(user: dict[str, Any]) -> bool:
    expires_at = parse_iso_date(user.get("account_expires_at"))
    return bool(expires_at and expires_at < today_utc())


def password_is_expired(user: dict[str, Any]) -> bool:
    expires_at = parse_iso_date(user.get("password_expires_at"))
    return bool(expires_at and expires_at < today_utc())


def public_user(username: str, user: dict[str, Any]) -> dict[str, Any]:
    total_seconds = float(user.get("total_session_seconds", 0.0) or 0.0)
    return {
        "username": username,
        "role": user.get("role", "user"),
        "full_name": user.get("full_name", ""),
        "email": user.get("email", ""),
        "account_expires_at": user.get("account_expires_at", ""),
        "password_changed_at": user.get("password_changed_at", ""),
        "password_expires_at": user.get("password_expires_at", ""),
        "last_login_at": user.get("last_login_at", ""),
        "session_count": int(user.get("session_count", 0) or 0),
        "total_session_seconds": total_seconds,
        "total_session_hours": round(total_seconds / 3600.0, 3),
        "modules_used": sorted({str(module) for module in user.get("modules_used", []) if str(module).strip()}),
    }


def public_users(users: dict[str, Any]) -> list[dict[str, Any]]:
    return [public_user(username, users[username]) for username in sorted(users)]


def ensure_data_files() -> None:
    DATA_ROOT.mkdir(exist_ok=True)
    REPORT_ROOT.mkdir(exist_ok=True)
    if not USERS_FILE.exists():
        USERS_FILE.write_text(
            json.dumps(
                {
                    "admin": make_user_record("admin123", "admin", full_name="Administrator", email="admin@example.com"),
                    "user": make_user_record("password123", "user", full_name="Default User", email="user@example.com"),
                },
                indent=2,
            ),
            encoding="utf-8",
        )
    if not METHOD_FILE.exists():
        METHOD_FILE.write_text(
            json.dumps(
                {
                    "method_name": "CEPA vehicle crossing methodology",
                    "soil_load_models": ["prism", "trap_door"],
                    "fatigue_constant": 1_000_000_000_000_000,
                    "fatigue_exponent": 3,
                    "notes": "Live load pressure uses Boussinesq point-load distribution and hoop bending uses CEPA-style Kb/Kz relationship.",
                },
                indent=2,
            ),
            encoding="utf-8",
        )


def read_users() -> dict[str, Any]:
    ensure_data_files()
    users = json.loads(USERS_FILE.read_text(encoding="utf-8"))
    changed = False
    for username, record in list(users.items()):
        normalized, record_changed = normalize_user_record(record)
        users[username] = normalized
        changed = changed or record_changed
    if changed:
        write_users(users)
    return users


def write_users(users: dict[str, Any]) -> None:
    ensure_data_files()
    USERS_FILE.write_text(json.dumps(users, indent=2), encoding="utf-8")


def start_user_session(users: dict[str, Any], username: str) -> str:
    user = users[username]
    session_id = secrets.token_hex(16)
    user["session_count"] = int(user.get("session_count", 0) or 0) + 1
    user["last_login_at"] = now_utc_iso()
    active_sessions = dict(user.get("active_sessions", {}))
    active_sessions[session_id] = {
        "started_at": user["last_login_at"],
        "last_seen_at": user["last_login_at"],
        "reported_seconds": 0.0,
        "modules_used": [],
    }
    user["active_sessions"] = active_sessions
    users[username] = user
    return session_id


def record_session_activity(
    users: dict[str, Any],
    username: str,
    session_id: str,
    elapsed_seconds: float,
    module: str = "",
    finished: bool = False,
) -> dict[str, Any]:
    if username not in users:
        raise ValueError("User was not found.")
    user = users[username]
    active_sessions = dict(user.get("active_sessions", {}))
    if session_id not in active_sessions:
        active_sessions[session_id] = {
            "started_at": now_utc_iso(),
            "last_seen_at": now_utc_iso(),
            "reported_seconds": 0.0,
            "modules_used": [],
        }
    session = dict(active_sessions[session_id])
    previous = float(session.get("reported_seconds", 0.0) or 0.0)
    reported = max(previous, max(0.0, float(elapsed_seconds or 0.0)))
    user["total_session_seconds"] = float(user.get("total_session_seconds", 0.0) or 0.0) + (reported - previous)
    session["reported_seconds"] = reported
    session["last_seen_at"] = now_utc_iso()
    module_name = str(module or "").strip()
    if module_name:
        modules = {str(item) for item in user.get("modules_used", []) if str(item).strip()}
        modules.add(module_name)
        user["modules_used"] = sorted(modules)
        session_modules = {str(item) for item in session.get("modules_used", []) if str(item).strip()}
        session_modules.add(module_name)
        session["modules_used"] = sorted(session_modules)
    if finished:
        active_sessions.pop(session_id, None)
    else:
        active_sessions[session_id] = session
    user["active_sessions"] = active_sessions
    users[username] = user
    return user


def session_is_active(user: dict[str, Any], session_id: str) -> bool:
    return bool(session_id and session_id in dict(user.get("active_sessions", {})))


def read_method_config() -> dict[str, Any]:
    ensure_data_files()
    return json.loads(METHOD_FILE.read_text(encoding="utf-8"))


def write_method_config(config: dict[str, Any]) -> None:
    ensure_data_files()
    METHOD_FILE.write_text(json.dumps(config, indent=2), encoding="utf-8")


def positive_number(data: dict[str, Any], key: str) -> float:
    value = float(data[key])
    if value <= 0:
        raise ValueError(f"{key} must be greater than zero")
    return value


def make_vehicle_loads(vehicle_data: dict[str, Any]) -> tuple[PointLoad, ...]:
    vehicle_type = str(vehicle_data["vehicle_type"])
    axle_count = int(vehicle_data["axle_count"])
    axle_width = positive_number(vehicle_data, "axle_width_in")
    axle_spacing = positive_number(vehicle_data, "axle_spacing_in")
    axle_loads = [float(load) for load in vehicle_data["axle_loads_lb"]]
    axle_widths = [float(width) for width in vehicle_data.get("axle_widths_in", [axle_width] * axle_count)]
    axle_spacings = [float(spacing) for spacing in vehicle_data.get("axle_spacings_in", [axle_spacing] * axle_count)]

    if axle_count <= 0:
        raise ValueError("axle_count must be greater than zero")
    if len(axle_loads) != axle_count:
        raise ValueError("axle_loads_lb length must match axle_count")
    if len(axle_widths) != axle_count:
        raise ValueError("axle_widths_in length must match axle_count")
    if len(axle_spacings) not in {axle_count, max(axle_count - 1, 0)}:
        raise ValueError("axle_spacings_in length must match axle_count or axle_count - 1")
    if any(load <= 0 for load in axle_loads):
        raise ValueError("each axle load must be greater than zero")
    if any(width <= 0 for width in axle_widths):
        raise ValueError("each axle width must be greater than zero")
    if any(spacing <= 0 for spacing in axle_spacings):
        raise ValueError("each axle spacing must be greater than zero")

    axle_positions = [0.0]
    for index in range(1, axle_count):
        axle_positions.append(axle_positions[-1] + axle_spacings[index - 1])
    origin = (axle_positions[0] + axle_positions[-1]) / 2.0
    loads: list[PointLoad] = []

    for index, axle_load in enumerate(axle_loads):
        x = axle_positions[index] - origin
        point_load = axle_load / 2.0
        y_offset = axle_widths[index] / 2.0
        prefix = "track" if vehicle_type == "track" else "axle"
        loads.append(PointLoad(f"{prefix}-{index + 1}-left", x, -y_offset, point_load))
        loads.append(PointLoad(f"{prefix}-{index + 1}-right", x, y_offset, point_load))

    return tuple(loads)


def case_from_gui_payload(payload: dict[str, Any]) -> tuple[CrossingCase, dict[str, Any]]:
    pipe_data = payload["pipe"]
    soil_data = payload["soil"]
    vehicle_data = payload["vehicle"]

    pipe = Pipe(
        outside_diameter_in=positive_number(pipe_data, "outside_diameter_in"),
        wall_thickness_in=positive_number(pipe_data, "wall_thickness_in"),
        mop_psig=float(pipe_data["maop_psig"]),
        smys_psi=positive_number(pipe_data, "smys_psi"),
        youngs_modulus_psi=positive_number(pipe_data, "youngs_modulus_psi"),
    )
    class_location = str(pipe_data.get("class_location", "1"))
    if class_location not in CSA_Z662_LOCATION_FACTORS:
        raise ValueError("class_location must be 1, 2, 3, or 4")
    design_factor = float(pipe_data.get("design_factor", DEFAULT_CSA_Z662_DESIGN_FACTOR))
    if design_factor <= 0 or design_factor > 1:
        raise ValueError("design_factor must be greater than zero and no greater than one")
    location_factor = CSA_Z662_LOCATION_FACTORS[class_location]
    hoop_limit_factor = design_factor * location_factor
    soil = Soil(
        cover_in=positive_number(soil_data, "cover_in"),
        unit_weight_pcf=positive_number(soil_data, "unit_weight_pcf"),
        modulus_reaction_psi=positive_number(soil_data, "modulus_reaction_psi"),
        kb=positive_number(soil_data, "kb"),
        kz=positive_number(soil_data, "kz"),
        load_model=str(soil_data.get("load_model", "prism")),
        friction_angle_deg=float(soil_data.get("friction_angle_deg", 30.0)),
    )
    soil_profile = str(soil_data.get("profile", "custom"))
    surface_layer_thickness = float(soil_data.get("surface_layer_thickness_in", 0.0))
    surface_layer_unit_weight = float(soil_data.get("surface_layer_unit_weight_pcf", 0.0))
    if surface_layer_thickness < 0:
        raise ValueError("surface_layer_thickness_in cannot be negative")
    if surface_layer_unit_weight < 0:
        raise ValueError("surface_layer_unit_weight_pcf cannot be negative")
    loads = make_vehicle_loads(vehicle_data)
    axle_count = int(vehicle_data["axle_count"])
    axle_spacings = [float(spacing) for spacing in vehicle_data.get("axle_spacings_in", [vehicle_data["axle_spacing_in"]] * axle_count)]
    axle_widths = [float(width) for width in vehicle_data.get("axle_widths_in", [vehicle_data["axle_width_in"]] * axle_count)]
    contact_areas = [float(area) for area in vehicle_data.get("contact_areas_in2", [vehicle_data["contact_area_in2"]] * axle_count)]
    if len(contact_areas) != axle_count:
        raise ValueError("contact_areas_in2 length must match axle_count")
    if len(axle_spacings) not in {axle_count, max(axle_count - 1, 0)}:
        raise ValueError("axle_spacings_in length must match axle_count or axle_count - 1")
    if any(area <= 0 for area in contact_areas):
        raise ValueError("each contact area must be greater than zero")
    axle_span = sum(axle_spacings[: max(axle_count - 1, 0)]) if axle_count > 1 else max(axle_spacings)
    envelope = max(axle_span, max(axle_widths), soil.cover_in * 2.0)
    scan_step = positive_number(vehicle_data, "scan_step_in")
    case = CrossingCase(
        pipe=pipe,
        soil=soil,
        vehicle=Vehicle(
            crossing_angle_deg=float(vehicle_data["crossing_angle_deg"]),
            impact_factor=positive_number(vehicle_data, "impact_factor"),
            loads=loads,
        ),
        scan=Scan(start_in=-envelope - soil.cover_in * 2.0, stop_in=envelope + soil.cover_in * 2.0, step_in=scan_step),
    )
    loaded_contact_area = sum(area * 2.0 for area in contact_areas)
    tire_widths = [float(width) for width in vehicle_data.get("tire_widths_in", [vehicle_data["tire_width_in"]])]
    if any(width <= 0 for width in tire_widths):
        raise ValueError("each tyre width must be greater than zero")
    track_lengths = [float(length) for length in vehicle_data.get("track_lengths_in", [])]
    if any(length <= 0 for length in track_lengths):
        raise ValueError("each track length must be greater than zero")
    vehicle_weight = sum(float(load) for load in vehicle_data["axle_loads_lb"])
    if str(vehicle_data.get("vehicle_type", "wheel")) == "track":
        tire_pressures = []
    else:
        tire_pressures = [float(pressure) for pressure in vehicle_data.get("tire_pressures_psi", [vehicle_data["tire_pressure_psi"]])]
        if any(pressure <= 0 for pressure in tire_pressures):
            raise ValueError("each tyre pressure must be greater than zero")
    metadata = {
        "vehicle_weight_lb": vehicle_weight,
        "point_load_count": len(loads),
        "average_point_load_lb": vehicle_weight / len(loads),
        "average_tire_width_in": sum(tire_widths) / len(tire_widths),
        "tire_widths_in": tire_widths,
        "contact_areas_in2": contact_areas,
        "average_contact_area_in2": sum(contact_areas) / len(contact_areas),
        "axle_widths_in": axle_widths,
        "average_axle_width_in": sum(axle_widths) / len(axle_widths),
        "axle_spacings_in": axle_spacings,
        "average_axle_spacing_in": sum(axle_spacings) / len(axle_spacings),
        "track_lengths_in": track_lengths,
        "average_tire_pressure_psi": sum(tire_pressures) / len(tire_pressures) if tire_pressures else None,
        "tire_pressures_psi": tire_pressures,
        "contact_pressure_psi": vehicle_weight / loaded_contact_area,
        "loads": [{"name": load.name, "x_in": load.x_in, "y_in": load.y_in, "load_lb": load.load_lb} for load in loads],
    }
    metadata["pipe"] = {
        "class_location": class_location,
        "design_factor": design_factor,
        "location_factor": location_factor,
        "hoop_limit_percent_smys": hoop_limit_factor * 100.0,
        "hoop_limit_psi": pipe.smys_psi * hoop_limit_factor,
    }
    metadata["soil"] = {
        "profile": soil_profile,
        "load_model": soil.load_model,
        "friction_angle_deg": soil.friction_angle_deg,
        "surface_layer_thickness_in": surface_layer_thickness,
        "surface_layer_unit_weight_pcf": surface_layer_unit_weight,
    }
    return case, metadata


def calculate_gui_payload(payload: dict[str, Any]) -> dict[str, Any]:
    case, metadata = case_from_gui_payload(payload)
    result = calculate_case(case)
    strain_data = payload.get("strain", {})
    bending_strain_microstrain = float(strain_data.get("bending_strain_microstrain", 0.0))
    if bending_strain_microstrain < 0:
        raise ValueError("bending_strain_microstrain cannot be negative")
    bending_stress_psi = case.pipe.youngs_modulus_psi * bending_strain_microstrain / 1_000_000.0
    for key in ("zero_pressure", "mop"):
        result[key]["pre_existing_bending_strain_microstrain"] = bending_strain_microstrain
        result[key]["pre_existing_bending_stress_psi"] = bending_stress_psi
        result[key]["assessment_stress_psi"] = result[key]["total_hoop_stress_psi"] + bending_stress_psi
        result[key]["assessment_percent_smys"] = 100.0 * result[key]["assessment_stress_psi"] / case.pipe.smys_psi
    result["pre_existing_bending_strain"] = {
        "bending_strain_microstrain": bending_strain_microstrain,
        "bending_stress_psi": bending_stress_psi,
        "basis": str(strain_data.get("basis", "absolute")),
    }
    result["vehicle"] = metadata
    return result


def calculate_corlas_payload(payload: dict[str, Any]) -> dict[str, Any]:
    geometry = payload.get("geometry", {})
    crack = payload.get("crack", {})
    material = payload.get("material", {})
    solver = payload.get("solver", {})

    diameter = positive_number(geometry, "outside_diameter_mm")
    wall_thickness = positive_number(geometry, "wall_thickness_mm")
    crack_depth = positive_number(crack, "depth_mm")
    crack_length = positive_number(crack, "length_mm")
    yield_strength = positive_number(material, "yield_strength_mpa")
    tensile_strength = positive_number(material, "tensile_strength_mpa")
    elastic_modulus = positive_number(material, "elastic_modulus_mpa")
    fracture_toughness = positive_number(material, "fracture_toughness_j")
    cvn = float(material.get("cvn_j", 0) or 0)
    charpy_area = float(material.get("charpy_area_in2", 0.124) or 0.124)
    toughness_method = str(material.get("fracture_toughness_method", "manual"))
    if toughness_method == "eq17_cvn":
        if cvn <= 0:
            raise ValueError("CVN must be greater than zero when Eq. 17 fracture toughness is selected")
        if charpy_area <= 0:
            raise ValueError("charpy_area_in2 must be greater than zero")
        fracture_toughness = 12.0 * cvn / charpy_area
    flow_coefficient = float(solver.get("flow_stress_coefficient", 0.5))
    pressure_step = positive_number(solver, "pressure_step_mpa")
    max_iterations = int(float(solver.get("max_iterations", 100000)))

    if wall_thickness >= diameter / 2:
        raise ValueError("wall_thickness_mm must be less than half the outside diameter")
    if crack_depth >= wall_thickness:
        raise ValueError("crack depth must be less than wall thickness")
    crack_location = str(crack.get("location", "external")).lower()
    crack_profile = str(crack.get("profile", "semi_elliptical")).lower()
    if crack_location not in {"external", "internal"}:
        raise ValueError("crack location must be external or internal")
    if crack_profile not in {"semi_elliptical", "parabolic", "rectangular"}:
        raise ValueError("crack profile must be semi_elliptical, parabolic, or rectangular")
    if crack_length <= 2 * crack_depth:
        raise ValueError("crack length should be greater than twice the crack depth for the semi-elliptical model")
    if tensile_strength <= yield_strength:
        raise ValueError("tensile strength must be greater than yield strength")
    if max_iterations <= 0:
        raise ValueError("max_iterations must be greater than zero")

    flow_stress = yield_strength + flow_coefficient * (tensile_strength - yield_strength)
    ratio = crack_length**2 / (diameter * wall_thickness)
    if ratio <= 50:
        folias_factor = math.sqrt(1 + 0.6275 * ratio - 0.003375 * ratio**2)
    else:
        folias_factor = 3.3 + 0.032 * ratio

    area_coefficients = {
        "semi_elliptical": math.pi / 4.0,
        "parabolic": 2.0 / 3.0,
        "rectangular": 1.0,
    }
    effective_area = area_coefficients[crack_profile] * crack_depth * crack_length
    reference_area = crack_length * wall_thickness
    area_ratio = effective_area / reference_area
    failure_stress = flow_stress * ((1 - area_ratio) / (1 - effective_area / (folias_factor * reference_area)))
    collapse_pressure = 2 * wall_thickness * failure_stress / diameter

    qf = (
        1.2581
        - 0.20589 * (crack_depth / crack_length)
        - 11.493 * (crack_depth / crack_length) ** 2
        + 29.586 * (crack_depth / crack_length) ** 3
        - 23.584 * (crack_depth / crack_length) ** 4
    )
    fsf = (
        (2 * wall_thickness / (math.pi * crack_depth))
        * math.tan(math.pi * crack_depth / (2 * wall_thickness))
        * (1 - 2 * crack_depth / crack_length)
        + 2 * crack_depth / crack_length
    )
    strain_hardening = -0.00546 + 0.556 * (yield_strength / tensile_strength) - 0.547 * (yield_strength / tensile_strength) ** 2
    plastic_strain_yield = 0.005 - yield_strength / elastic_modulus
    if strain_hardening <= 0:
        raise ValueError("Calculated strain hardening exponent is not positive. Check yield and tensile strength.")
    if plastic_strain_yield <= 0:
        raise ValueError("Calculated plastic strain at yield is not positive. Check yield strength and elastic modulus.")

    pressure = 0.0
    iterations = 0
    hoop_stress = 0.0
    local_stress = 0.0
    plastic_strain = 0.0
    shih_hutchinson = 0.0
    elastic_j = 0.0
    plastic_j = 0.0
    total_j = 0.0
    stopped_by = "fracture"

    while True:
        hoop_stress = pressure * diameter / (2 * wall_thickness)
        crack_face_pressure = pressure * math.pi * crack_depth / (4 * wall_thickness) if crack_location == "internal" else 0.0
        local_stress = (hoop_stress + crack_face_pressure) * (
            (1 - math.pi * crack_depth / (4 * wall_thickness * folias_factor)) / (1 - math.pi * crack_depth / (4 * wall_thickness))
        )
        plastic_strain = plastic_strain_yield * (local_stress / yield_strength) ** (1 / strain_hardening)
        shih_hutchinson = ((3.85 * math.sqrt(1 / strain_hardening) * (1 - strain_hardening)) + math.pi * strain_hardening) * (
            1 + strain_hardening
        )
        elastic_j = qf * fsf * crack_depth * ((local_stress**2) * math.pi / elastic_modulus)
        plastic_j = qf * fsf * crack_depth * shih_hutchinson * plastic_strain * local_stress
        total_j = elastic_j + plastic_j
        if total_j >= fracture_toughness:
            stopped_by = "fracture"
            break
        pressure += pressure_step
        iterations += 1
        if pressure > collapse_pressure:
            stopped_by = "collapse"
            break
        if iterations >= max_iterations:
            stopped_by = "iteration_limit"
            break

    failure_pressure = min(pressure, collapse_pressure)
    controlling_mode = "Fracture" if pressure <= collapse_pressure else "Plastic collapse"
    psi_per_mpa = 145.037737796858
    return {
        "inputs": {
            "outside_diameter_mm": diameter,
            "wall_thickness_mm": wall_thickness,
            "crack_depth_mm": crack_depth,
            "crack_length_mm": crack_length,
            "yield_strength_mpa": yield_strength,
            "tensile_strength_mpa": tensile_strength,
            "elastic_modulus_mpa": elastic_modulus,
            "fracture_toughness_j": fracture_toughness,
            "fracture_toughness_method": toughness_method,
            "cvn_j": cvn,
            "charpy_area_in2": charpy_area,
            "crack_location": crack_location,
            "crack_profile": crack_profile,
            "flow_stress_coefficient": flow_coefficient,
            "pressure_step_mpa": pressure_step,
        },
        "outputs": {
            "fracture_pressure_mpa": pressure,
            "collapse_pressure_mpa": collapse_pressure,
            "failure_pressure_mpa": failure_pressure,
            "fracture_pressure_psi": pressure * psi_per_mpa,
            "collapse_pressure_psi": collapse_pressure * psi_per_mpa,
            "failure_pressure_psi": failure_pressure * psi_per_mpa,
            "controlling_mode": controlling_mode,
            "stopped_by": stopped_by,
            "iterations": iterations,
        },
        "intermediate": {
            "flow_stress_mpa": flow_stress,
            "folias_ratio": ratio,
            "folias_factor_m": folias_factor,
            "effective_flaw_area_mm2": effective_area,
            "effective_area_coefficient": area_coefficients[crack_profile],
            "reference_area_mm2": reference_area,
            "area_ratio": area_ratio,
            "failure_stress_mpa": failure_stress,
            "qf": qf,
            "fsf": fsf,
            "strain_hardening_n": strain_hardening,
            "plastic_strain_yield": plastic_strain_yield,
            "hoop_stress_mpa": hoop_stress,
            "local_stress_mpa": local_stress,
            "plastic_strain": plastic_strain,
            "shih_hutchinson_f3": shih_hutchinson,
            "elastic_j": elastic_j,
            "plastic_j": plastic_j,
            "total_j": total_j,
        },
    }


def calculate_annex_k_eca_payload(payload: dict[str, Any]) -> dict[str, Any]:
    geometry = payload.get("geometry", {})
    material = payload.get("material", {})
    loads = payload.get("loads", {})
    flaw = payload.get("flaw", {})
    assessment = payload.get("assessment", {})

    diameter = positive_number(geometry, "outside_diameter_mm")
    wall_thickness = positive_number(geometry, "wall_thickness_mm")
    smys = positive_number(material, "smys_mpa")
    weld_yield = positive_number(material, "weld_yield_strength_mpa")
    base_yield = positive_number(material, "base_yield_strength_mpa")
    elastic_modulus = positive_number(material, "elastic_modulus_mpa")
    poisson = float(material.get("poisson_ratio", 0.3))
    thermal_coeff = positive_number(material, "thermal_coefficient_per_c")
    toughness = positive_number(material, "kmat_mpa_sqrt_m")
    pressure = float(loads.get("pressure_mpa", 0.0))
    delta_t = float(loads.get("temperature_change_c", 0.0))
    bending_moment = float(loads.get("bending_moment_kn_m", 0.0))
    misalignment = float(flaw.get("misalignment_mm", 0.0))
    measured_height = positive_number(flaw, "measured_height_mm")
    measured_length = positive_number(flaw, "measured_length_mm")
    height_allowance = float(flaw.get("height_nde_allowance_mm", 0.5))
    length_allowance = float(flaw.get("length_nde_allowance_mm", 2.0))
    residual_factor = float(assessment.get("residual_stress_factor", 0.6))
    service = str(assessment.get("service_type", "liquid")).lower()
    longitudinal_strain = float(assessment.get("longitudinal_strain_percent", 0.0))

    if not (0 <= poisson < 0.5):
        raise ValueError("poisson_ratio must be between 0 and 0.5")
    if pressure < 0 or bending_moment < 0 or misalignment < 0:
        raise ValueError("pressure, bending moment, and misalignment cannot be negative")
    if height_allowance < 0 or length_allowance < 0:
        raise ValueError("NDE allowances cannot be negative")
    if service not in {"liquid", "gas"}:
        raise ValueError("service_type must be liquid or gas")

    effective_height = measured_height + height_allowance
    effective_length = measured_length + length_allowance
    scf = 1 + (3 * misalignment / wall_thickness) * (1 / (1 - poisson**2))
    hoop_stress = pressure * diameter / (2 * wall_thickness)
    inside_diameter = diameter - 2 * wall_thickness
    section_modulus = (math.pi / 32.0) * (diameter**4 - inside_diameter**4) / diameter
    bending_stress = bending_moment * 1_000_000.0 / section_modulus if section_modulus > 0 else 0.0
    thermal_stress = elastic_modulus * thermal_coeff * delta_t
    pressure_axial_stress = poisson * hoop_stress
    axial_stress = pressure_axial_stress + thermal_stress + bending_stress
    mismatch_ratio = weld_yield / base_yield
    load_ratio = axial_stress * scf / base_yield
    fracture_ratio = (1.12 * (axial_stress * scf + weld_yield * residual_factor) * math.sqrt(math.pi * (effective_height / 1000.0))) / toughness
    fad_boundary = (1 - 0.14 * load_ratio**2) * (0.3 + 0.7 * math.exp(-0.6 * load_ratio**6))

    height_limit_fraction = 0.25 if service == "liquid" else 0.50
    height_limit = height_limit_fraction * wall_thickness
    length_limit = 0.10 * math.pi * diameter
    gateway_issues = []
    if effective_height > height_limit:
        gateway_issues.append(f"Effective flaw height exceeds {height_limit_fraction:.0%} wall-thickness gateway.")
    if effective_length > length_limit:
        gateway_issues.append("Effective flaw length exceeds 10% of pipe circumference gateway.")

    if load_ratio > 1.2:
        status = "REJECT - PLASTIC COLLAPSE"
    elif fracture_ratio > fad_boundary:
        status = "REJECT - FRACTURE BOUNDARY EXCEEDED"
    elif gateway_issues:
        status = "CONDITIONALLY ACCEPTABLE - GATEWAY REVIEW REQUIRED"
    else:
        status = "STATUS: ACCEPTABLE"

    if "REJECT" in status:
        disposition = "Repair Required"
    elif gateway_issues:
        disposition = "Conditionally Acceptable"
    elif longitudinal_strain > 0.5 or mismatch_ratio < 1.0:
        disposition = "Conditionally Acceptable"
    else:
        disposition = "Acceptable"

    if longitudinal_strain > 0.5 or mismatch_ratio < 1.0:
        assessment_level = "Level 3 - Advanced nonlinear FEA recommended"
    elif abs(axial_stress) <= 0.5 * smys:
        assessment_level = "Level 1 - Conservative screening"
    else:
        assessment_level = "Level 2 - Elastic-plastic FAD"

    return {
        "inputs": {
            "outside_diameter_mm": diameter,
            "wall_thickness_mm": wall_thickness,
            "smys_mpa": smys,
            "weld_yield_strength_mpa": weld_yield,
            "base_yield_strength_mpa": base_yield,
            "elastic_modulus_mpa": elastic_modulus,
            "poisson_ratio": poisson,
            "thermal_coefficient_per_c": thermal_coeff,
            "kmat_mpa_sqrt_m": toughness,
            "pressure_mpa": pressure,
            "temperature_change_c": delta_t,
            "bending_moment_kn_m": bending_moment,
            "misalignment_mm": misalignment,
            "measured_height_mm": measured_height,
            "measured_length_mm": measured_length,
            "height_nde_allowance_mm": height_allowance,
            "length_nde_allowance_mm": length_allowance,
            "service_type": service,
        },
        "outputs": {
            "status": status,
            "disposition": disposition,
            "assessment_level": assessment_level,
            "gateway_issues": gateway_issues,
            "effective_flaw_height_mm": effective_height,
            "effective_flaw_length_mm": effective_length,
            "scf": scf,
            "hoop_stress_mpa": hoop_stress,
            "pressure_axial_stress_mpa": pressure_axial_stress,
            "thermal_stress_mpa": thermal_stress,
            "bending_stress_mpa": bending_stress,
            "restrained_axial_stress_mpa": axial_stress,
            "section_modulus_mm3": section_modulus,
            "strength_mismatch_ratio": mismatch_ratio,
            "load_ratio_lr": load_ratio,
            "fracture_ratio_kr": fracture_ratio,
            "fad_boundary": fad_boundary,
            "height_gateway_limit_mm": height_limit,
            "length_gateway_limit_mm": length_limit,
        },
    }


def calculate_dent_assessment_payload(payload: dict[str, Any]) -> dict[str, Any]:
    pipe = payload.get("pipe", {})
    dent = payload.get("dent", {})
    simulation = payload.get("simulation", {})
    od = positive_number(pipe, "outside_diameter_in")
    wall_thickness = positive_number(pipe, "wall_thickness_in")
    r0 = od / 2.0
    r1_mean = float(dent.get("circumferential_radius_in", 0.0))
    r2_mean = float(dent.get("longitudinal_radius_in", 0.0))
    depth_mean = positive_number(dent, "depth_in")
    length_mean = positive_number(dent, "length_in")
    error_percent = float(simulation.get("measurement_error_fraction", 0.10))
    strain_limit = positive_number(simulation, "strain_limit")
    num_simulations = int(float(simulation.get("num_simulations", 100000)))
    seed = int(float(simulation.get("seed", 8675309)))

    if r1_mean >= 0 or r2_mean >= 0:
        raise ValueError("Dent radii should be negative for inward dents.")
    if error_percent < 0:
        raise ValueError("measurement_error_fraction cannot be negative")
    if num_simulations < 100:
        raise ValueError("num_simulations must be at least 100")

    rng = random.Random(seed)
    peak_strains: list[float] = []
    for _ in range(num_simulations):
        r1 = rng.gauss(r1_mean, abs(r1_mean) * error_percent)
        r2 = rng.gauss(r2_mean, abs(r2_mean) * error_percent)
        depth = rng.gauss(depth_mean, depth_mean * error_percent)
        length = rng.gauss(length_mean, length_mean * error_percent)
        if r1 >= -0.1:
            r1 = -0.1
        if r2 >= -0.1:
            r2 = -0.1
        if depth < 0.01:
            depth = 0.01
        if length < 0.1:
            length = 0.1
        eps_1 = (wall_thickness / 2.0) * ((1.0 / r0) - (1.0 / r1))
        eps_2 = wall_thickness / (2.0 * r2)
        eps_3 = 0.5 * (depth / length) ** 2
        eps_i = math.sqrt(eps_1**2 - eps_1 * (eps_2 + eps_3) + (eps_2 + eps_3) ** 2)
        eps_o = math.sqrt(eps_1**2 + eps_1 * (eps_2 - eps_3) + (eps_2 - eps_3) ** 2)
        peak_strains.append(max(eps_i, eps_o))

    peak_strains.sort()
    mean_strain = sum(peak_strains) / num_simulations
    variance = sum((value - mean_strain) ** 2 for value in peak_strains) / max(num_simulations - 1, 1)
    std_strain = math.sqrt(variance)
    p95_strain = peak_strains[min(num_simulations - 1, max(0, math.ceil(num_simulations * 0.95) - 1))]
    exceed_count = sum(1 for value in peak_strains if value > strain_limit)
    probability_exceedance = exceed_count / num_simulations
    status = "REPAIR REQUIRED" if p95_strain > strain_limit else "ACCEPTABLE"
    disposition = "Repair Required" if status == "REPAIR REQUIRED" else "Acceptable"
    depth_percent_od = depth_mean / od
    framework_notes = []
    if depth_percent_od > 0.06:
        framework_notes.append("Dent depth exceeds 6% OD screening threshold; advanced FFS review recommended.")
    if probability_exceedance > 0.05:
        framework_notes.append("More than 5% of simulations exceed the strain limit.")
    if not framework_notes:
        framework_notes.append("Statistical strain screening is within the selected limit.")

    sample_count = 101
    distribution = []
    for index in range(sample_count):
        position = min(num_simulations - 1, round(index * (num_simulations - 1) / (sample_count - 1)))
        distribution.append({"peak_strain": peak_strains[position], "cumulative_probability": index / (sample_count - 1)})

    return {
        "inputs": {
            "outside_diameter_in": od,
            "wall_thickness_in": wall_thickness,
            "circumferential_radius_in": r1_mean,
            "longitudinal_radius_in": r2_mean,
            "depth_in": depth_mean,
            "length_in": length_mean,
            "measurement_error_fraction": error_percent,
            "strain_limit": strain_limit,
            "num_simulations": num_simulations,
            "seed": seed,
        },
        "outputs": {
            "mean_peak_strain": mean_strain,
            "std_dev_strain": std_strain,
            "p95_strain": p95_strain,
            "probability_exceedance": probability_exceedance,
            "status": status,
            "disposition": disposition,
            "depth_percent_od": depth_percent_od,
            "exceed_count": exceed_count,
            "framework_notes": framework_notes,
            "distribution": distribution,
        },
    }


def calculate_modified_b31g_payload(payload: dict[str, Any]) -> dict[str, Any]:
    pipe = payload.get("pipe", {})
    defect = payload.get("defect", {})
    assessment = payload.get("assessment", {})
    diameter = positive_number(pipe, "outside_diameter_mm")
    wall_thickness = positive_number(pipe, "wall_thickness_mm")
    smys = positive_number(pipe, "smys_mpa")
    smts = positive_number(pipe, "smts_mpa")
    maop = float(pipe.get("maop_mpa", 0.0) or 0.0)
    depth = positive_number(defect, "depth_mm")
    length = positive_number(defect, "length_mm")
    assessment_factor = float(assessment.get("assessment_factor", 0.72) or 0.72)
    cap_flow_stress = bool(assessment.get("cap_flow_stress_to_smts", True))
    if maop < 0:
        raise ValueError("maop_mpa cannot be negative")
    if depth >= wall_thickness:
        raise ValueError("defect depth must be less than wall thickness")
    if assessment_factor <= 0:
        raise ValueError("assessment_factor must be greater than zero")

    z = length**2 / (diameter * wall_thickness)
    if z <= 50.0:
        folias = math.sqrt(max(1.0 + 0.6275 * z - 0.003375 * z**2, 1.0))
        folias_equation = "sqrt(1 + 0.6275z - 0.003375z^2), z <= 50"
    else:
        folias = 3.3 + 0.032 * z
        folias_equation = "3.3 + 0.032z, z > 50"

    flow_stress_uncapped = smys + 69.0
    flow_stress = min(flow_stress_uncapped, smts) if cap_flow_stress else flow_stress_uncapped
    depth_ratio = depth / wall_thickness
    numerator = 1.0 - 0.85 * depth_ratio
    denominator = 1.0 - (0.85 * depth_ratio / folias)
    if numerator <= 0 or denominator <= 0:
        raise ValueError("defect geometry is outside the Modified B31G equation range")

    failure_stress = flow_stress * numerator / denominator
    failure_pressure = (2.0 * wall_thickness * failure_stress) / diameter
    allowable_pressure = failure_pressure * assessment_factor
    pressure_ratio = maop / allowable_pressure if allowable_pressure > 0 else float("inf")
    remaining_pressure_factor = allowable_pressure / maop if maop > 0 else float("inf")
    status = "PASS" if maop <= allowable_pressure else "FAIL"
    disposition = (
        "MAOP is less than or equal to the factored Modified B31G pressure."
        if status == "PASS"
        else "MAOP exceeds the factored Modified B31G pressure. Repair, pressure reduction, or further assessment is required."
    )

    return {
        "inputs": {
            "outside_diameter_mm": diameter,
            "wall_thickness_mm": wall_thickness,
            "smys_mpa": smys,
            "smts_mpa": smts,
            "maop_mpa": maop,
            "defect_depth_mm": depth,
            "defect_length_mm": length,
            "assessment_factor": assessment_factor,
            "cap_flow_stress_to_smts": cap_flow_stress,
        },
        "outputs": {
            "status": status,
            "disposition": disposition,
            "depth_ratio": depth_ratio,
            "z_parameter": z,
            "folias_factor": folias,
            "folias_equation": folias_equation,
            "flow_stress_mpa": flow_stress,
            "flow_stress_uncapped_mpa": flow_stress_uncapped,
            "failure_stress_mpa": failure_stress,
            "failure_pressure_mpa": failure_pressure,
            "allowable_pressure_mpa": allowable_pressure,
            "maop_to_allowable_ratio": pressure_ratio,
            "remaining_pressure_factor": remaining_pressure_factor,
        },
    }


def calculate_rstreng_payload(payload: dict[str, Any]) -> dict[str, Any]:
    pipe = payload.get("pipe", {})
    assessment = payload.get("assessment", {})
    profile = payload.get("profile", [])
    diameter = positive_number(pipe, "outside_diameter_mm")
    wall_thickness = positive_number(pipe, "wall_thickness_mm")
    smys = positive_number(pipe, "smys_mpa")
    smts = positive_number(pipe, "smts_mpa")
    maop = float(pipe.get("maop_mpa", 0.0) or 0.0)
    assessment_factor = float(assessment.get("assessment_factor", 0.72) or 0.72)
    cap_flow_stress = bool(assessment.get("cap_flow_stress_to_smts", True))
    if maop < 0:
        raise ValueError("maop_mpa cannot be negative")
    if assessment_factor <= 0:
        raise ValueError("assessment_factor must be greater than zero")
    if not isinstance(profile, list) or len(profile) < 2:
        raise ValueError("profile must include at least two station/depth points")

    points: list[tuple[float, float]] = []
    for item in profile:
        station = float(item.get("station_mm"))
        depth = float(item.get("depth_mm"))
        if station < 0:
            raise ValueError("profile station_mm cannot be negative")
        if depth < 0:
            raise ValueError("profile depth_mm cannot be negative")
        if depth >= wall_thickness:
            raise ValueError("profile depth_mm values must be less than wall thickness")
        points.append((station, depth))
    points.sort(key=lambda item: item[0])
    if len({station for station, _depth in points}) != len(points):
        raise ValueError("profile station_mm values must be unique")
    if points[-1][0] <= points[0][0]:
        raise ValueError("profile must have a positive total length")

    flow_stress_uncapped = smys + 69.0
    flow_stress = min(flow_stress_uncapped, smts) if cap_flow_stress else flow_stress_uncapped

    def folias_factor(length: float) -> tuple[float, str, float]:
        z_parameter = length**2 / (diameter * wall_thickness)
        if z_parameter <= 50.0:
            return (
                math.sqrt(max(1.0 + 0.6275 * z_parameter - 0.003375 * z_parameter**2, 1.0)),
                "sqrt(1 + 0.6275z - 0.003375z^2), z <= 50",
                z_parameter,
            )
        return 3.3 + 0.032 * z_parameter, "3.3 + 0.032z, z > 50", z_parameter

    def metal_loss_area(start_index: int, end_index: int) -> float:
        area = 0.0
        for index in range(start_index, end_index):
            x1, d1 = points[index]
            x2, d2 = points[index + 1]
            area += (d1 + d2) * 0.5 * (x2 - x1)
        return area

    controlling: dict[str, Any] | None = None
    for start_index in range(len(points) - 1):
        for end_index in range(start_index + 1, len(points)):
            start_station = points[start_index][0]
            end_station = points[end_index][0]
            length = end_station - start_station
            if length <= 0:
                continue
            area = metal_loss_area(start_index, end_index)
            original_area = length * wall_thickness
            area_ratio = min(area / original_area, 0.999999)
            folias, folias_equation, z_parameter = folias_factor(length)
            numerator = 1.0 - area_ratio
            denominator = 1.0 - (area_ratio / folias)
            if numerator <= 0 or denominator <= 0:
                failure_stress = 0.0
            else:
                failure_stress = flow_stress * numerator / denominator
            failure_pressure = max((2.0 * wall_thickness * failure_stress) / diameter, 0.001)
            candidate = {
                "start_station_mm": start_station,
                "end_station_mm": end_station,
                "length_mm": length,
                "effective_area_mm2": area,
                "original_area_mm2": original_area,
                "area_ratio": area_ratio,
                "average_depth_mm": area / length,
                "max_depth_mm": max(depth for _station, depth in points[start_index : end_index + 1]),
                "folias_factor": folias,
                "folias_equation": folias_equation,
                "z_parameter": z_parameter,
                "failure_stress_mpa": failure_stress,
                "failure_pressure_mpa": failure_pressure,
            }
            if controlling is None or candidate["failure_pressure_mpa"] < controlling["failure_pressure_mpa"]:
                controlling = candidate

    if controlling is None:
        raise ValueError("profile did not contain a valid assessment segment")

    allowable_pressure = controlling["failure_pressure_mpa"] * assessment_factor
    pressure_ratio = maop / allowable_pressure if allowable_pressure > 0 else float("inf")
    remaining_pressure_factor = allowable_pressure / maop if maop > 0 else float("inf")
    status = "PASS" if maop <= allowable_pressure else "FAIL"
    disposition = (
        "MAOP is less than or equal to the factored RSTRENG effective-area pressure."
        if status == "PASS"
        else "MAOP exceeds the factored RSTRENG effective-area pressure. Repair, pressure reduction, or detailed review is required."
    )

    return {
        "inputs": {
            "outside_diameter_mm": diameter,
            "wall_thickness_mm": wall_thickness,
            "smys_mpa": smys,
            "smts_mpa": smts,
            "maop_mpa": maop,
            "assessment_factor": assessment_factor,
            "cap_flow_stress_to_smts": cap_flow_stress,
            "profile": [{"station_mm": station, "depth_mm": depth} for station, depth in points],
        },
        "outputs": {
            "status": status,
            "disposition": disposition,
            "flow_stress_mpa": flow_stress,
            "flow_stress_uncapped_mpa": flow_stress_uncapped,
            "allowable_pressure_mpa": allowable_pressure,
            "maop_to_allowable_ratio": pressure_ratio,
            "remaining_pressure_factor": remaining_pressure_factor,
            "controlling_segment": controlling,
        },
    }


def calculate_scc_colony_payload(payload: dict[str, Any]) -> dict[str, Any]:
    pipe = payload.get("pipe", {})
    colony = payload.get("colony", {})
    fatigue = payload.get("fatigue", {})
    diameter = positive_number(pipe, "outside_diameter_mm")
    wall_thickness = positive_number(pipe, "wall_thickness_mm")
    maop = positive_number(pipe, "maop_mpa")
    smys = positive_number(pipe, "smys_mpa")
    smts = positive_number(pipe, "smts_mpa")
    toughness = positive_number(pipe, "fracture_toughness_mpa_sqrt_m")
    assessment_factor = float(pipe.get("assessment_factor", 0.72) or 0.72)
    if assessment_factor <= 0:
        raise ValueError("assessment_factor must be greater than zero")

    depths = [float(item) for item in colony.get("depths_mm", [])]
    lengths = [float(item) for item in colony.get("lengths_mm", [])]
    spacings = [float(item) for item in colony.get("axial_spacings_mm", [])]
    if not depths or len(depths) != len(lengths):
        raise ValueError("depths_mm and lengths_mm must contain the same number of cracks")
    if len(depths) < 2:
        raise ValueError("At least two cracks are required for SCC colony assessment")
    if len(spacings) not in {0, len(depths) - 1}:
        raise ValueError("axial_spacings_mm must be empty or have one fewer value than the crack count")
    if not spacings:
        spacings = [0.0] * (len(depths) - 1)
    if any(depth <= 0 or depth >= wall_thickness for depth in depths):
        raise ValueError("Each crack depth must be greater than zero and less than wall thickness")
    if any(length <= 0 for length in lengths):
        raise ValueError("Each crack length must be greater than zero")
    if any(spacing < 0 for spacing in spacings):
        raise ValueError("Crack spacings cannot be negative")

    orientation = str(colony.get("orientation", "axial"))
    y_factor = float(colony.get("geometry_factor", 1.12) or 1.12)
    if y_factor <= 0:
        raise ValueError("geometry_factor must be greater than zero")

    sqrt_dt = math.sqrt(diameter * wall_thickness)
    normalized_spacings = [spacing / sqrt_dt for spacing in spacings]
    proximity_terms = [math.exp(-0.85 * spacing_norm) for spacing_norm in normalized_spacings]
    interaction_index = sum(proximity_terms)
    severity = max(depths) / wall_thickness
    interaction_factor = 1.0 + min(0.65, 0.12 * interaction_index + 0.35 * severity * interaction_index)
    equivalent_depth = min(max(depths) * interaction_factor, wall_thickness * 0.95)
    close_spans = [index for index, spacing_norm in enumerate(normalized_spacings) if spacing_norm <= 1.5]
    if close_spans:
        first = min(close_spans)
        last = max(close_spans) + 1
        equivalent_length = sum(lengths[first : last + 1]) + sum(spacings[first:last])
    else:
        governing = max(range(len(depths)), key=lambda index: depths[index] * math.sqrt(lengths[index]))
        first = governing
        last = governing
        equivalent_length = lengths[governing]

    z = equivalent_length**2 / (diameter * wall_thickness)
    if z <= 50.0:
        folias = math.sqrt(max(1.0 + 0.6275 * z - 0.003375 * z**2, 1.0))
        folias_equation = "sqrt(1 + 0.6275z - 0.003375z^2), z <= 50"
    else:
        folias = 3.3 + 0.032 * z
        folias_equation = "3.3 + 0.032z, z > 50"

    flow_stress = min(smys + 69.0, smts)
    depth_ratio = equivalent_depth / wall_thickness
    collapse_denominator = 1.0 - (depth_ratio / folias)
    collapse_stress = flow_stress * (1.0 - depth_ratio) / collapse_denominator if collapse_denominator > 0 else 0.0
    crack_depth_m = equivalent_depth / 1000.0
    fracture_stress = toughness / (y_factor * math.sqrt(math.pi * crack_depth_m))
    pressure_stress_factor = diameter / (2.0 * wall_thickness) if orientation == "axial" else diameter / (4.0 * wall_thickness)
    collapse_pressure = max(collapse_stress / pressure_stress_factor, 0.001)
    fracture_pressure = max(fracture_stress / pressure_stress_factor, 0.001)
    failure_pressure = min(collapse_pressure, fracture_pressure)
    allowable_pressure = failure_pressure * assessment_factor
    maop_ratio = maop / allowable_pressure if allowable_pressure > 0 else float("inf")
    hoop_stress = maop * pressure_stress_factor
    kr_at_maop = y_factor * hoop_stress * math.sqrt(math.pi * crack_depth_m) / toughness
    status = "ACCEPTABLE" if maop <= allowable_pressure else "ACTION REQUIRED"
    if interaction_factor < 1.12:
        colony_class = "Weak interaction"
    elif interaction_factor < 1.35:
        colony_class = "Moderate interaction"
    else:
        colony_class = "Strong interaction"
    disposition = (
        "MAOP is less than or equal to the factored SCC colony failure pressure."
        if status == "ACCEPTABLE"
        else "MAOP exceeds the factored SCC colony failure pressure. Consider pressure reduction, repair, replacement, or advanced ECA."
    )

    pressure_range = max(0.0, float(fatigue.get("pressure_range_mpa", 0.0) or 0.0))
    paris_c = float(fatigue.get("paris_c", 1e-12) or 1e-12)
    paris_m = float(fatigue.get("paris_m", 3.0) or 3.0)
    critical_depth = min(0.8 * wall_thickness, wall_thickness - 0.001)
    remaining_cycles = None
    crack_growth_status = "Not checked"
    if pressure_range > 0 and paris_c > 0 and paris_m > 0 and equivalent_depth < critical_depth:
        delta_stress = pressure_range * pressure_stress_factor
        delta_k = y_factor * delta_stress * math.sqrt(math.pi * crack_depth_m)
        growth_per_cycle = paris_c * (delta_k**paris_m)
        remaining_cycles = (critical_depth - equivalent_depth) / growth_per_cycle if growth_per_cycle > 0 else None
        crack_growth_status = "Estimated" if remaining_cycles is not None else "Not available"

    return {
        "inputs": {
            "outside_diameter_mm": diameter,
            "wall_thickness_mm": wall_thickness,
            "maop_mpa": maop,
            "smys_mpa": smys,
            "smts_mpa": smts,
            "fracture_toughness_mpa_sqrt_m": toughness,
            "assessment_factor": assessment_factor,
            "orientation": orientation,
            "geometry_factor": y_factor,
            "depths_mm": depths,
            "lengths_mm": lengths,
            "axial_spacings_mm": spacings,
        },
        "outputs": {
            "status": status,
            "disposition": disposition,
            "colony_class": colony_class,
            "crack_count": len(depths),
            "first_interacting_crack": first + 1,
            "last_interacting_crack": last + 1,
            "interaction_factor": interaction_factor,
            "equivalent_depth_mm": equivalent_depth,
            "equivalent_length_mm": equivalent_length,
            "depth_ratio": depth_ratio,
            "folias_factor": folias,
            "folias_equation": folias_equation,
            "z_parameter": z,
            "flow_stress_mpa": flow_stress,
            "collapse_pressure_mpa": collapse_pressure,
            "fracture_pressure_mpa": fracture_pressure,
            "failure_pressure_mpa": failure_pressure,
            "allowable_pressure_mpa": allowable_pressure,
            "maop_to_allowable_ratio": maop_ratio,
            "kr_at_maop": kr_at_maop,
            "remaining_cycles": remaining_cycles,
            "crack_growth_status": crack_growth_status,
        },
    }


def calculate_crack_growth_payload(payload: dict[str, Any]) -> dict[str, Any]:
    crack = payload.get("crack", {})
    loading = payload.get("loading", {})
    assessment = payload.get("assessment", {})

    initial_crack = positive_number(crack, "initial_crack_mm")
    critical_crack = positive_number(crack, "critical_crack_mm")
    if critical_crack <= initial_crack:
        raise ValueError("critical_crack_mm must be greater than initial_crack_mm")

    stress_range = positive_number(loading, "stress_range_mpa")
    geometry_factor = float(loading.get("geometry_factor", 1.12) or 1.12)
    if geometry_factor <= 0:
        raise ValueError("geometry_factor must be greater than zero")
    threshold_delta_k = max(0.0, float(loading.get("threshold_delta_k_mpa_sqrt_m", 0.0) or 0.0))

    paris_c = positive_number(assessment, "paris_c")
    paris_m = positive_number(assessment, "paris_m")
    increment = float(assessment.get("increment_mm", 0.001) or 0.001)
    if increment <= 0:
        raise ValueError("increment_mm must be greater than zero")
    applied_cycles = max(0.0, float(assessment.get("applied_cycles", 0.0) or 0.0))
    life_factor = float(assessment.get("life_factor", 1.0) or 1.0)
    if life_factor <= 0:
        raise ValueError("life_factor must be greater than zero")

    def delta_k_for_crack(crack_mm: float) -> float:
        return geometry_factor * stress_range * math.sqrt(math.pi * (crack_mm / 1000.0))

    def growth_rate_for_crack(crack_mm: float) -> float:
        delta_k = delta_k_for_crack(crack_mm)
        if delta_k <= threshold_delta_k:
            return 0.0
        return paris_c * ((delta_k - threshold_delta_k) ** paris_m)

    crack_size = initial_crack
    cycles = 0.0
    max_steps = 2_000_000
    steps = 0
    while crack_size < critical_crack:
        step = min(increment, critical_crack - crack_size)
        midpoint = crack_size + step / 2.0
        growth_rate = growth_rate_for_crack(midpoint)
        if growth_rate <= 0:
            cycles = float("inf")
            break
        cycles += step / growth_rate
        crack_size += step
        steps += 1
        if steps > max_steps:
            raise ValueError("Crack growth integration did not converge; increase increment_mm")

    factored_cycles = cycles * life_factor if math.isfinite(cycles) else float("inf")
    damage_ratio = applied_cycles / factored_cycles if factored_cycles > 0 and math.isfinite(factored_cycles) else 0.0
    remaining_cycles = max(factored_cycles - applied_cycles, 0.0) if math.isfinite(factored_cycles) else None
    status = "PASS" if applied_cycles <= factored_cycles else "FAIL"
    if not math.isfinite(cycles):
        status = "NO GROWTH"
        disposition = "Stress-intensity range is at or below the threshold; Paris-law crack growth is not predicted."
    elif status == "PASS":
        disposition = "Applied cycles are within the factored Paris-law crack-growth life."
    else:
        disposition = "Applied cycles exceed the factored Paris-law crack-growth life. Reduce cyclic loading, repair, replace, or complete detailed ECA."

    return {
        "inputs": {
            "initial_crack_mm": initial_crack,
            "critical_crack_mm": critical_crack,
            "stress_range_mpa": stress_range,
            "geometry_factor": geometry_factor,
            "threshold_delta_k_mpa_sqrt_m": threshold_delta_k,
            "paris_c": paris_c,
            "paris_m": paris_m,
            "increment_mm": increment,
            "applied_cycles": applied_cycles,
            "life_factor": life_factor,
        },
        "outputs": {
            "status": status,
            "disposition": disposition,
            "estimated_cycles": cycles if math.isfinite(cycles) else None,
            "factored_cycles": factored_cycles if math.isfinite(factored_cycles) else None,
            "applied_cycles": applied_cycles,
            "remaining_cycles": remaining_cycles,
            "damage_ratio": damage_ratio,
            "initial_delta_k_mpa_sqrt_m": delta_k_for_crack(initial_crack),
            "critical_delta_k_mpa_sqrt_m": delta_k_for_crack(critical_crack),
            "initial_growth_rate_mm_per_cycle": growth_rate_for_crack(initial_crack),
            "critical_growth_rate_mm_per_cycle": growth_rate_for_crack(critical_crack),
            "integration_steps": steps,
        },
    }


def calculate_ili_screening_payload(payload: dict[str, Any]) -> dict[str, Any]:
    pipe = payload.get("pipe", {})
    features = payload.get("features", {})
    criteria = payload.get("criteria", {})
    fatigue = payload.get("fatigue", {})
    risk = payload.get("risk", {})

    diameter = positive_number(pipe, "outside_diameter_mm")
    wall_thickness = positive_number(pipe, "wall_thickness_mm")
    maop = positive_number(pipe, "maop_mpa")
    smys = positive_number(pipe, "smys_mpa")
    assessment_factor = float(pipe.get("assessment_factor", 0.72) or 0.72)
    if assessment_factor <= 0:
        raise ValueError("assessment_factor must be greater than zero")

    repair_ratio = float(criteria.get("repair_pressure_ratio", 1.0) or 1.0)
    monitor_ratio = float(criteria.get("monitor_pressure_ratio", 1.25) or 1.25)
    depth_watch_percent = float(criteria.get("depth_watch_percent", 50.0) or 50.0)
    if repair_ratio <= 0 or monitor_ratio <= 0 or depth_watch_percent <= 0:
        raise ValueError("ILI screening criteria must be greater than zero")
    primary_method = str(criteria.get("primary_method", "modified_b31g") or "modified_b31g")
    feature_methods = {str(key): str(value) for key, value in dict(criteria.get("feature_methods", {}) or {}).items()}
    selected_methods = [str(item) for item in criteria.get("screening_methods", ["modified_b31g"]) if str(item).strip()]
    if primary_method and primary_method not in selected_methods:
        selected_methods.insert(0, primary_method)
    if not selected_methods:
        selected_methods = ["modified_b31g"]
    class_location = str(risk.get("class_location", "1") or "1")
    prediction_years = max(0.0, float(risk.get("prediction_years", 5.0) or 5.0))
    annual_growth_percent = max(0.0, float(risk.get("annual_growth_percent", 0.0) or 0.0))
    location_factor = CSA_Z662_LOCATION_FACTORS.get(class_location, 1.0)
    fatigue_enabled = bool(fatigue.get("enabled", False))
    stress_range = max(0.0, float(fatigue.get("stress_range_mpa", 0.0) or 0.0))
    bending_strain_percent = max(0.0, float(fatigue.get("bending_strain_percent", 0.0) or 0.0))
    cycles_per_year = max(0.0, float(fatigue.get("cycles_per_year", 0.0) or 0.0))
    applied_cycles = max(0.0, float(fatigue.get("applied_cycles", 0.0) or 0.0))
    paris_c = float(fatigue.get("paris_c", 1e-12) or 1e-12)
    paris_m = float(fatigue.get("paris_m", 3.0) or 3.0)

    ids = [str(item).strip() for item in features.get("ids", [])]
    types = [str(item).strip() or "metal_loss" for item in features.get("types", [])]
    depths = [float(item) for item in features.get("depths_percent", [])]
    lengths = [float(item) for item in features.get("lengths_mm", [])]
    clocks = [str(item).strip() for item in features.get("clock_positions", [])]
    distances = [float(item) for item in features.get("distances_m", [])]
    reported_pressures = [float(item) if item not in (None, "") else 0.0 for item in features.get("reported_failure_pressures_mpa", [])]

    count = len(ids)
    if count == 0:
        raise ValueError("At least one ILI feature is required")
    for name, values in {
        "types": types,
        "depths_percent": depths,
        "lengths_mm": lengths,
        "clock_positions": clocks,
        "distances_m": distances,
    }.items():
        if len(values) != count:
            raise ValueError(f"{name} must contain one value per ILI feature")
    if reported_pressures and len(reported_pressures) != count:
        raise ValueError("reported_failure_pressures_mpa must be empty or contain one value per ILI feature")
    if not reported_pressures:
        reported_pressures = [0.0] * count

    flow_stress = smys + 69.0
    ranked = []
    immediate = high = monitor = acceptable = 0
    conservative_method_counts: dict[str, int] = {}
    predicted_failures = 0
    for index, feature_id in enumerate(ids):
        depth_percent = depths[index]
        length = lengths[index]
        if depth_percent <= 0 or depth_percent >= 100:
            raise ValueError("Each depth percentage must be greater than zero and less than 100")
        if length <= 0:
            raise ValueError("Each feature length must be greater than zero")
        depth_ratio = depth_percent / 100.0
        z = length**2 / (diameter * wall_thickness)
        folias = math.sqrt(max(1.0 + 0.6275 * z - 0.003375 * z**2, 1.0)) if z <= 50 else 3.3 + 0.032 * z
        def pressure_from_stress(stress: float) -> float:
            return max((2.0 * wall_thickness * stress) / diameter, 0.001)

        def modified_b31g_pressure(depth: float) -> float:
            denominator = 1.0 - depth / folias
            stress = flow_stress * (1.0 - depth) / denominator if denominator > 0 else 0.0
            return pressure_from_stress(stress)

        def asme_b31g_pressure(depth: float) -> float:
            denominator = 1.0 - ((2.0 / 3.0) * depth / folias)
            stress = smys * (1.0 - (2.0 / 3.0) * depth) / denominator if denominator > 0 else 0.0
            return pressure_from_stress(stress)

        def rstreng_simplified_pressure(depth: float) -> float:
            effective_depth = min(depth * 0.85, 0.95)
            denominator = 1.0 - effective_depth / folias
            stress = flow_stress * (1.0 - effective_depth) / denominator if denominator > 0 else 0.0
            return pressure_from_stress(stress)

        calculated_pressure = modified_b31g_pressure(depth_ratio)
        reported_pressure = reported_pressures[index]
        method_results: list[dict[str, Any]] = []
        method_labels = {
            "modified_b31g": "Modified B31.G",
            "asme_b31g": "ASME B31G",
            "rstreng_simplified": "RSTRENG simplified",
            "corlas": "CorLAS crack-like flaw",
            "scc_colony": "SCC crack colony",
            "reported_pressure": "ILI reported pressure",
            "crack_fracture": "Crack/SCC fracture screen",
        }
        toughness = max(60.0, float(criteria.get("fracture_toughness_mpa_sqrt_m", 95.0) or 95.0))
        crack_depth_m = max(depth_ratio * wall_thickness / 1000.0, 1e-6)
        requested_method = feature_methods.get(feature_id)
        feature_selected_methods = list(selected_methods)
        if requested_method and requested_method not in feature_selected_methods:
            feature_selected_methods.append(requested_method)
        for method in feature_selected_methods:
            pressure = None
            if method == "modified_b31g":
                pressure = modified_b31g_pressure(depth_ratio)
            elif method == "asme_b31g":
                pressure = asme_b31g_pressure(depth_ratio)
            elif method == "rstreng_simplified":
                pressure = rstreng_simplified_pressure(depth_ratio)
            elif method == "reported_pressure" and reported_pressure > 0:
                pressure = reported_pressure
            elif method == "corlas":
                fracture_stress = toughness / (1.12 * math.sqrt(math.pi * crack_depth_m))
                collapse_pressure = modified_b31g_pressure(depth_ratio)
                pressure = min(pressure_from_stress(fracture_stress), collapse_pressure)
            elif method == "scc_colony":
                interaction_factor = 1.0 + (0.18 if types[index].lower() in {"crack", "crack_like", "scc"} else 0.08)
                equivalent_depth = min(depth_ratio * interaction_factor, 0.95)
                fracture_stress = toughness / (1.12 * math.sqrt(math.pi * max(equivalent_depth * wall_thickness / 1000.0, 1e-6)))
                pressure = min(pressure_from_stress(fracture_stress), modified_b31g_pressure(equivalent_depth))
            elif method == "crack_fracture":
                fracture_stress = toughness / (1.12 * math.sqrt(math.pi * crack_depth_m))
                pressure = pressure_from_stress(fracture_stress)
            if pressure is not None:
                method_results.append(
                    {
                        "method": method,
                        "label": method_labels.get(method, method),
                        "failure_pressure_mpa": pressure,
                        "allowable_pressure_mpa": pressure * assessment_factor,
                        "maop_to_allowable_ratio": maop / (pressure * assessment_factor) if pressure > 0 else float("inf"),
                    }
                )
        if not method_results:
            method_results.append(
                {
                    "method": "modified_b31g",
                    "label": "Modified B31.G",
                    "failure_pressure_mpa": calculated_pressure,
                    "allowable_pressure_mpa": calculated_pressure * assessment_factor,
                    "maop_to_allowable_ratio": maop / (calculated_pressure * assessment_factor),
                }
            )
        conservative = min(method_results, key=lambda item: item["allowable_pressure_mpa"])
        conservative_method_counts[conservative["label"]] = conservative_method_counts.get(conservative["label"], 0) + 1
        calculation_method = next((item for item in method_results if item["method"] == requested_method), conservative)
        governing_pressure = calculation_method["failure_pressure_mpa"]
        allowable_pressure = governing_pressure * assessment_factor
        pressure_ratio = maop / allowable_pressure if allowable_pressure > 0 else float("inf")
        future_depth_percent = min(depth_percent + annual_growth_percent * prediction_years, 99.0)
        future_depth_ratio = future_depth_percent / 100.0
        future_pressure = min(
            modified_b31g_pressure(future_depth_ratio),
            conservative["failure_pressure_mpa"] if annual_growth_percent <= 0 else max(modified_b31g_pressure(future_depth_ratio), 0.001),
        )
        future_ratio = maop / (future_pressure * assessment_factor) if future_pressure > 0 else float("inf")
        predicted_failure_years = None
        if annual_growth_percent > 0:
            for year_index in range(1, int(math.ceil(prediction_years)) + 1):
                trial_depth = min((depth_percent + annual_growth_percent * year_index) / 100.0, 0.99)
                trial_pressure = modified_b31g_pressure(trial_depth)
                if maop / (trial_pressure * assessment_factor) >= 1.0:
                    predicted_failure_years = float(year_index)
                    break
        if predicted_failure_years is not None:
            predicted_failures += 1
        fatigue_life_years = None
        fatigue_life_cycles = None
        fatigue_damage_ratio = 0.0
        if fatigue_enabled and stress_range > 0 and paris_c > 0 and paris_m > 0:
            effective_stress = stress_range * (1.0 + bending_strain_percent / 2.0)
            crack_depth_m = max(depth_ratio * wall_thickness / 1000.0, 1e-6)
            delta_k = 1.12 * effective_stress * math.sqrt(math.pi * crack_depth_m)
            growth_rate = paris_c * (delta_k**paris_m)
            remaining_depth_mm = max(0.8 * wall_thickness - depth_ratio * wall_thickness, 0.0)
            if growth_rate > 0 and remaining_depth_mm > 0:
                fatigue_life_cycles = remaining_depth_mm / growth_rate
                fatigue_life_years = fatigue_life_cycles / cycles_per_year if cycles_per_year > 0 else None
                fatigue_damage_ratio = applied_cycles / fatigue_life_cycles if fatigue_life_cycles > 0 else 0.0
        depth_ratio_to_watch = depth_percent / depth_watch_percent
        length_ratio = length / math.sqrt(diameter * wall_thickness)
        type_factor = 1.25 if types[index].lower() in {"crack", "crack_like", "scc"} else 1.0
        class_factor = {"1": 1.0, "2": 1.12, "3": 1.28, "4": 1.45}.get(class_location, 1.0)
        fatigue_factor = 1.15 if fatigue_life_years is not None and fatigue_life_years <= prediction_years else 1.0
        future_factor = 1.25 if predicted_failure_years is not None else 1.0
        severity_score = max(pressure_ratio, depth_ratio_to_watch, future_ratio, min(length_ratio / 2.0, 1.5)) * type_factor * class_factor * fatigue_factor * future_factor
        if pressure_ratio >= repair_ratio:
            priority = "Immediate action"
            recommended_action = "Repair, replace, pressure-reduce, or complete detailed assessment before continued operation."
            immediate += 1
        elif pressure_ratio >= monitor_ratio or severity_score >= 1.25:
            priority = "High priority"
            recommended_action = "Schedule engineering review and confirm sizing, interaction, and excavation priority."
            high += 1
        elif depth_percent >= depth_watch_percent or severity_score >= 0.85:
            priority = "Monitor"
            recommended_action = "Track in the next ILI run and review growth or interaction assumptions."
            monitor += 1
        else:
            priority = "Acceptable"
            recommended_action = "Retain in feature register and continue normal integrity monitoring."
            acceptable += 1

        ranked.append(
            {
                "rank": 0,
                "feature_id": feature_id,
                "feature_type": types[index],
                "distance_m": distances[index],
                "clock_position": clocks[index],
                "depth_percent": depth_percent,
                "length_mm": length,
                "folias_factor": folias,
                "calculated_failure_pressure_mpa": calculated_pressure,
                "reported_failure_pressure_mpa": reported_pressure if reported_pressure > 0 else None,
                "governing_failure_pressure_mpa": governing_pressure,
                "allowable_pressure_mpa": allowable_pressure,
                "maop_to_allowable_ratio": pressure_ratio,
                "method_results": method_results,
                "conservative_method": conservative["label"],
                "calculation_method": calculation_method["label"],
                "calculation_method_key": calculation_method["method"],
                "future_depth_percent": future_depth_percent,
                "future_maop_to_allowable_ratio": future_ratio,
                "predicted_failure_years": predicted_failure_years,
                "fatigue_life_cycles": fatigue_life_cycles,
                "fatigue_life_years": fatigue_life_years,
                "fatigue_damage_ratio": fatigue_damage_ratio,
                "severity_score": severity_score,
                "risk_class": "High" if severity_score >= 1.5 else "Medium" if severity_score >= 0.95 else "Low",
                "priority": priority,
                "recommended_action": recommended_action,
            }
        )

    ranked.sort(key=lambda item: (-item["severity_score"], item["distance_m"], item["feature_id"]))
    for rank, item in enumerate(ranked, start=1):
        item["rank"] = rank

    status = "ACTION REQUIRED" if immediate else "REVIEW REQUIRED" if high else "MONITOR" if monitor else "ACCEPTABLE"
    most_conservative_method = max(conservative_method_counts.items(), key=lambda item: item[1])[0] if conservative_method_counts else "-"
    highest_risk_feature = ranked[0]["feature_id"] if ranked else "-"
    risk_class = "High" if immediate or predicted_failures else "Medium" if high or monitor else "Low"
    disposition = (
        f"{immediate} feature(s) require immediate action, {high} high-priority review, {monitor} monitoring, and {acceptable} acceptable. "
        f"Most conservative method overall: {most_conservative_method}."
    )
    return {
        "inputs": {
            "outside_diameter_mm": diameter,
            "wall_thickness_mm": wall_thickness,
            "maop_mpa": maop,
            "smys_mpa": smys,
            "assessment_factor": assessment_factor,
            "repair_pressure_ratio": repair_ratio,
            "monitor_pressure_ratio": monitor_ratio,
            "depth_watch_percent": depth_watch_percent,
            "screening_methods": selected_methods,
            "primary_method": primary_method,
            "feature_methods": feature_methods,
            "class_location": class_location,
            "prediction_years": prediction_years,
            "annual_growth_percent": annual_growth_percent,
            "fatigue_enabled": fatigue_enabled,
        },
        "outputs": {
            "status": status,
            "disposition": disposition,
            "feature_count": count,
            "immediate_count": immediate,
            "high_count": high,
            "monitor_count": monitor,
            "acceptable_count": acceptable,
            "most_conservative_method": most_conservative_method,
            "highest_risk_feature": highest_risk_feature,
            "predicted_failure_count": predicted_failures,
            "risk_class": risk_class,
            "class_location_factor": location_factor,
            "ranked_features": ranked,
        },
    }


def normalize_ili_import_header(value: Any) -> str:
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in str(value or "")).strip("_")


def value_from_ili_import_row(row: list[Any], headers: list[str], candidates: list[str], fallback_index: int) -> str:
    for candidate in candidates:
        if candidate in headers:
            index = headers.index(candidate)
            if index < len(row) and row[index] not in (None, ""):
                return str(row[index]).strip()
    return str(row[fallback_index]).strip() if fallback_index < len(row) and row[fallback_index] is not None else ""


def feature_lists_from_rows(rows: list[list[Any]]) -> dict[str, list[str]]:
    clean_rows = [[cell for cell in row] for row in rows if any(str(cell or "").strip() for cell in row)]
    if not clean_rows:
        raise ValueError("The uploaded ILI feature file is empty.")
    first = [normalize_ili_import_header(cell) for cell in clean_rows[0]]
    has_header = any(item in {"id", "feature_id", "feature", "type", "feature_type", "depth", "depth_percent"} for item in first)
    headers = first if has_header else []
    data_rows = clean_rows[1:] if has_header else clean_rows
    imported = {
        "ids": [],
        "types": [],
        "depths_percent": [],
        "lengths_mm": [],
        "clock_positions": [],
        "distances_m": [],
        "reported_failure_pressures_mpa": [],
    }
    for row in data_rows:
        if len(row) < 6:
            continue
        imported["ids"].append(value_from_ili_import_row(row, headers, ["id", "feature_id", "feature"], 0))
        imported["types"].append(value_from_ili_import_row(row, headers, ["type", "feature_type", "anomaly_type"], 1))
        imported["depths_percent"].append(value_from_ili_import_row(row, headers, ["depth", "depth_percent", "depth_pct", "depth_wall"], 2))
        imported["lengths_mm"].append(value_from_ili_import_row(row, headers, ["length", "length_mm", "axial_length"], 3))
        imported["clock_positions"].append(value_from_ili_import_row(row, headers, ["clock", "clock_position"], 4))
        imported["distances_m"].append(value_from_ili_import_row(row, headers, ["distance", "distance_m", "odometer", "chainage"], 5))
        imported["reported_failure_pressures_mpa"].append(
            value_from_ili_import_row(row, headers, ["pressure", "failure_pressure", "failure_pressure_mpa", "pfail"], 6) or "0"
        )
    if not imported["ids"]:
        raise ValueError("No feature rows were found. Expected columns: id, type, depth, length, clock, distance, pressure.")
    return imported


def parse_ili_feature_file(filename: str, content: bytes) -> dict[str, list[str]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        try:
            import openpyxl
        except Exception as exc:  # pragma: no cover - environment dependent
            raise ValueError("Excel import requires openpyxl on the backend.") from exc
        workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        return feature_lists_from_rows(rows)
    text = content.decode("utf-8-sig")
    sample = text[:1024]
    delimiter = "\t" if "\t" in sample else ";" if ";" in sample and sample.count(";") > sample.count(",") else ","
    rows = [row for row in csv.reader(StringIO(text), delimiter=delimiter)]
    return feature_lists_from_rows(rows)


RAW_ILI_SCHEMAS: dict[str, tuple[tuple[str, tuple[str, ...]], ...]] = {
    "mfl": (
        ("feature_id", ("feature_id", "feature", "id", "anomaly_id")),
        ("distance_m", ("distance_m", "distance", "odometer", "chainage")),
        ("clock_position", ("clock_position", "clock", "clock_pos")),
        ("axial_offset_mm", ("axial_offset_mm", "axial_mm", "x_mm", "axial")),
        ("circumferential_offset_mm", ("circumferential_offset_mm", "circumferential_mm", "y_mm", "circumferential")),
        ("depth_percent", ("depth_percent", "depth_pct", "depth", "metal_loss_percent")),
    ),
    "crack": (
        ("feature_id", ("feature_id", "feature", "id", "colony_id")),
        ("distance_m", ("distance_m", "distance", "odometer", "chainage")),
        ("clock_position", ("clock_position", "clock", "clock_pos")),
        ("axial_offset_mm", ("axial_offset_mm", "axial_mm", "x_mm", "axial")),
        ("circumferential_offset_mm", ("circumferential_offset_mm", "circumferential_mm", "y_mm", "circumferential")),
        ("depth_mm", ("depth_mm", "depth", "crack_depth_mm")),
        ("opening_mm", ("opening_mm", "opening", "cod_mm", "crack_opening_mm")),
        ("orientation_deg", ("orientation_deg", "orientation", "angle_deg")),
        ("crack_id", ("crack_id", "segment_id", "branch_id")),
        ("anomaly_type", ("anomaly_type", "type", "feature_type")),
    ),
    "caliper": (
        ("feature_id", ("feature_id", "feature", "id", "anomaly_id")),
        ("distance_m", ("distance_m", "distance", "odometer", "chainage")),
        ("clock_position", ("clock_position", "clock", "clock_pos")),
        ("axial_offset_mm", ("axial_offset_mm", "axial_mm", "x_mm", "axial")),
        ("circumferential_offset_mm", ("circumferential_offset_mm", "circumferential_mm", "y_mm", "circumferential")),
        ("radial_deformation_mm", ("radial_deformation_mm", "radial_mm", "deformation_mm", "dent_depth_mm")),
    ),
}


def rows_from_uploaded_table(filename: str, content: bytes) -> list[list[Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        try:
            import openpyxl
        except Exception as exc:  # pragma: no cover - environment dependent
            raise ValueError("Excel import requires openpyxl on the backend.") from exc
        workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        return [list(row) for row in workbook.active.iter_rows(values_only=True)]
    text = content.decode("utf-8-sig")
    sample = text[:2048]
    delimiter = "\t" if "\t" in sample else ";" if sample.count(";") > sample.count(",") else ","
    return [row for row in csv.reader(StringIO(text), delimiter=delimiter)]


def parse_ili_raw_file(filename: str, content: bytes, tool_type: str) -> list[dict[str, Any]]:
    tool = tool_type.strip().lower()
    if tool not in RAW_ILI_SCHEMAS:
        raise ValueError("Raw ILI tool type must be mfl, crack, or caliper.")
    rows = [row for row in rows_from_uploaded_table(filename, content) if any(str(cell or "").strip() for cell in row)]
    if not rows:
        raise ValueError(f"The uploaded {tool.upper()} file is empty.")
    headers = [normalize_ili_import_header(cell) for cell in rows[0]]
    schema = RAW_ILI_SCHEMAS[tool]
    has_header = any(candidate in headers for _, candidates in schema for candidate in candidates)
    if not has_header:
        headers = [name for name, _ in schema]
        data_rows = rows
    else:
        data_rows = rows[1:]

    samples: list[dict[str, Any]] = []
    for row_index, row in enumerate(data_rows, start=1):
        sample: dict[str, Any] = {}
        for fallback_index, (name, candidates) in enumerate(schema):
            raw_value = value_from_ili_import_row(row, headers, list(candidates), fallback_index)
            if name in {"feature_id", "clock_position", "crack_id", "anomaly_type"}:
                sample[name] = str(raw_value or "").strip()
            else:
                sample[name] = float(raw_value or 0.0)
        sample["feature_id"] = sample["feature_id"] or f"{tool.upper()}-001"
        if tool == "mfl" and not (0 < sample["depth_percent"] < 100):
            raise ValueError(f"MFL row {row_index} depth_percent must be greater than zero and less than 100.")
        if tool == "crack" and sample["depth_mm"] <= 0:
            raise ValueError(f"Crack row {row_index} depth_mm must be greater than zero.")
        if tool == "caliper" and sample["radial_deformation_mm"] == 0:
            raise ValueError(f"Caliper row {row_index} radial_deformation_mm cannot be zero.")
        samples.append(sample)
    if not samples:
        raise ValueError(f"No valid {tool.upper()} rows were found.")
    return samples


def parse_multipart_file(content_type: str, body: bytes) -> tuple[str, bytes]:
    marker = "boundary="
    if marker not in content_type:
        raise ValueError("Missing multipart boundary.")
    boundary = content_type.split(marker, 1)[1].strip().strip('"')
    boundary_bytes = f"--{boundary}".encode("utf-8")
    for part in body.split(boundary_bytes):
        if b"Content-Disposition" not in part or b"filename=" not in part:
            continue
        header, _, data = part.partition(b"\r\n\r\n")
        disposition = header.decode("utf-8", errors="ignore")
        filename = "ili-features.csv"
        if "filename=" in disposition:
            filename = disposition.split("filename=", 1)[1].split(";", 1)[0].strip().strip('"')
        data = data.removesuffix(b"\r\n")
        data = data.removesuffix(b"--")
        data = data.removesuffix(b"\r\n")
        return filename, data
    raise ValueError("No uploaded file was found.")


def calculate_interacting_anomalies_payload(payload: dict[str, Any]) -> dict[str, Any]:
    pipe = payload.get("pipe", {})
    loading = payload.get("loading", {})
    uncertainty = payload.get("uncertainty", {})
    mesh = payload.get("mesh", {})
    anomaly_items = payload.get("anomalies", [])
    if len(anomaly_items) < 2:
        raise ValueError("At least two anomalies are required for interaction assessment.")

    diameter = positive_number(pipe, "outside_diameter_mm")
    wall_thickness = positive_number(pipe, "wall_thickness_mm")
    maop = positive_number(pipe, "maop_mpa")
    smys = positive_number(pipe, "smys_mpa")
    smts = positive_number(pipe, "smts_mpa")
    elastic_modulus = positive_number(pipe, "elastic_modulus_mpa")
    toughness = positive_number(pipe, "fracture_toughness_mpa_sqrt_m")
    model_length_factor = float(pipe.get("model_length_factor", 8.0) or 8.0)
    if model_length_factor < 6 or model_length_factor > 10:
        raise ValueError("model_length_factor should be between 6D and 10D for this framework.")

    depth_tolerance = max(0.0, float(uncertainty.get("depth_tolerance_mm", 0.0) or 0.0))
    length_tolerance = max(0.0, float(uncertainty.get("length_tolerance_mm", 0.0) or 0.0))
    width_tolerance = max(0.0, float(uncertainty.get("width_tolerance_mm", 0.0) or 0.0))
    sizing_case = str(uncertainty.get("case", "nominal"))
    uncertainty_scale = {"nominal": 0.0, "conservative": 1.0, "probabilistic": 0.5}.get(sizing_case, 0.0)

    flow_stress = min(smys + 69.0, smts)
    flow_pressure = 2.0 * wall_thickness * flow_stress / diameter
    sqrt_dt = math.sqrt(diameter * wall_thickness)

    def anomaly_record(raw: dict[str, Any], index: int) -> dict[str, Any]:
        length = positive_number(raw, "length_mm") + uncertainty_scale * length_tolerance
        width = positive_number(raw, "width_mm") + uncertainty_scale * width_tolerance
        depth = positive_number(raw, "depth_mm") + uncertainty_scale * depth_tolerance
        if depth >= wall_thickness:
            raise ValueError(f"Anomaly {index} depth must be less than wall thickness after uncertainty adjustment.")
        return {
            "id": f"A{index}",
            "type": str(raw.get("type", "metal_loss")),
            "surface": str(raw.get("surface", "external")),
            "axial_location_mm": float(raw.get("axial_location_mm", 0.0) or 0.0),
            "clock_position_deg": float(raw.get("clock_position_deg", 0.0) or 0.0) % 360.0,
            "length_mm": length,
            "width_mm": width,
            "depth_mm": depth,
            "orientation_deg": float(raw.get("orientation_deg", 0.0) or 0.0),
            "reported_length_mm": positive_number(raw, "length_mm"),
            "reported_width_mm": positive_number(raw, "width_mm"),
            "reported_depth_mm": positive_number(raw, "depth_mm"),
        }

    anomalies = [anomaly_record(raw, idx + 1) for idx, raw in enumerate(anomaly_items[:2])]

    def modified_b31g_failure_pressure(anomaly: dict[str, Any]) -> float:
        depth = anomaly["depth_mm"]
        length = anomaly["length_mm"]
        depth_ratio = depth / wall_thickness
        z_value = length**2 / (diameter * wall_thickness)
        if z_value <= 50:
            folias = math.sqrt(1 + 0.6275 * z_value - 0.003375 * z_value**2)
        else:
            folias = 3.3 + 0.032 * z_value
        numerator = 1 - 0.85 * depth_ratio
        denominator = 1 - 0.85 * depth / (wall_thickness * folias)
        if numerator <= 0 or denominator <= 0:
            return max(0.05 * flow_pressure, 0.001)
        failure_stress = flow_stress * numerator / denominator
        return max(2.0 * wall_thickness * failure_stress / diameter, 0.001)

    def crack_failure_pressure(anomaly: dict[str, Any]) -> float:
        crack_depth_m = anomaly["depth_mm"] / 1000.0
        orientation = math.radians(anomaly["orientation_deg"])
        y_factor = 1.12
        fracture_stress = toughness / (y_factor * math.sqrt(math.pi * crack_depth_m))
        stress_per_mpa_pressure = (diameter / (2.0 * wall_thickness)) * (math.cos(orientation) ** 2) + (
            diameter / (4.0 * wall_thickness)
        ) * (math.sin(orientation) ** 2)
        fracture_pressure = fracture_stress / max(stress_per_mpa_pressure, 1e-9)
        plastic_pressure = flow_pressure * max(0.15, 1.0 - 0.35 * anomaly["depth_mm"] / wall_thickness)
        return max(min(fracture_pressure, plastic_pressure), 0.001)

    def dent_failure_pressure(anomaly: dict[str, Any]) -> float:
        depth_od = anomaly["depth_mm"] / diameter
        depth_t = anomaly["depth_mm"] / wall_thickness
        return max(flow_pressure / (1.0 + 8.0 * depth_od + 0.55 * depth_t), 0.001)

    def isolated_failure_pressure(anomaly: dict[str, Any]) -> float:
        kind = anomaly["type"]
        if kind == "metal_loss":
            return modified_b31g_failure_pressure(anomaly)
        if kind == "crack":
            return crack_failure_pressure(anomaly)
        if kind == "dent":
            return dent_failure_pressure(anomaly)
        if kind == "weld":
            return 0.9 * flow_pressure
        return min(modified_b31g_failure_pressure(anomaly), crack_failure_pressure(anomaly))

    for anomaly in anomalies:
        anomaly["failure_pressure_mpa"] = isolated_failure_pressure(anomaly)
        anomaly["depth_ratio"] = anomaly["depth_mm"] / wall_thickness
        anomaly["length_parameter"] = anomaly["length_mm"] / sqrt_dt

    a1, a2 = anomalies
    delta_theta = abs(a2["clock_position_deg"] - a1["clock_position_deg"])
    delta_theta = min(delta_theta, 360.0 - delta_theta)
    axial_spacing = max(0.0, abs(a2["axial_location_mm"] - a1["axial_location_mm"]) - (a1["length_mm"] + a2["length_mm"]) / 2.0)
    circum_spacing = max(
        0.0,
        (diameter / 2.0) * math.radians(delta_theta) - (a1["width_mm"] + a2["width_mm"]) / 2.0,
    )
    lambda_x = axial_spacing / sqrt_dt
    lambda_theta = circum_spacing / sqrt_dt
    spacing_index = math.sqrt(lambda_x**2 + lambda_theta**2)
    proximity = math.exp(-0.55 * spacing_index**2)
    if axial_spacing == 0 and circum_spacing == 0:
        proximity = 1.0

    types = {a1["type"], a2["type"]}
    mixed_multiplier = 1.0
    if len(types) > 1:
        mixed_multiplier += 0.12
    if "crack" in types and "metal_loss" in types:
        mixed_multiplier += 0.13
    if "dent" in types and "crack" in types:
        mixed_multiplier += 0.15

    average_severity = (a1["depth_ratio"] + a2["depth_ratio"]) / 2.0
    overlap_bonus = 0.06 if axial_spacing == 0 or circum_spacing == 0 else 0.0
    uncertainty_multiplier = {"nominal": 1.0, "conservative": 1.08, "probabilistic": 1.04}.get(sizing_case, 1.0)
    interaction_factor = 1.0 + proximity * (0.08 + 0.55 * average_severity) * mixed_multiplier * uncertainty_multiplier + overlap_bonus
    interaction_factor = max(1.0, min(interaction_factor, 2.5))

    secondary_stress = abs(float(loading.get("secondary_stress_mpa", 0.0) or 0.0))
    residual_fraction = max(0.0, float(loading.get("residual_stress_fraction", 0.0) or 0.0))
    load_modifier = 1.0 + 0.35 * secondary_stress / max(smys, 1e-9) + 0.18 * residual_fraction
    weakest_isolated = min(anomaly["failure_pressure_mpa"] for anomaly in anomalies)
    combined_failure_pressure = weakest_isolated / (interaction_factor * load_modifier)
    safety_factor = combined_failure_pressure / maop

    if interaction_factor < 1.05:
        interaction_classification = "Negligible"
        interaction_response = "Separate assessment may be acceptable if validation and other criteria are satisfied."
    elif interaction_factor < 1.25:
        interaction_classification = "Moderate"
        interaction_response = "Evaluate as an interacting cluster for reassessment and monitoring."
    else:
        interaction_classification = "Strong"
        interaction_response = "Treat as a single complex defect and prioritize remediation or pressure management."

    if safety_factor > 1.5:
        safety_category = "Acceptable margin"
        recommended_response = "Continued operation may be acceptable with routine monitoring after validation."
        status = "ACCEPTABLE"
    elif safety_factor > 1.25:
        safety_category = "Marginal"
        recommended_response = "Review pressure transients, increase monitoring, and plan conservative reassessment or remediation."
        status = "MARGINAL"
    else:
        safety_category = "Critical"
        recommended_response = "Consider immediate pressure reduction, repair, replacement, or shutdown under operator procedures."
        status = "ACTION REQUIRED"

    pressure_range = max(0.0, float(loading.get("pressure_range_mpa", 0.0) or 0.0))
    crack_anomalies = [a for a in anomalies if a["type"] in {"crack", "mixed"}]
    k_max = 0.0
    remaining_cycles = None
    if crack_anomalies:
        controlling_crack = max(crack_anomalies, key=lambda item: item["depth_mm"])
        crack_depth_m = controlling_crack["depth_mm"] / 1000.0
        orientation = math.radians(controlling_crack["orientation_deg"])
        stress_per_mpa_pressure = (diameter / (2.0 * wall_thickness)) * (math.cos(orientation) ** 2) + (
            diameter / (4.0 * wall_thickness)
        ) * (math.sin(orientation) ** 2)
        k_max = 1.12 * maop * stress_per_mpa_pressure * math.sqrt(math.pi * crack_depth_m)
        if pressure_range > 0:
            delta_k = 1.12 * pressure_range * stress_per_mpa_pressure * math.sqrt(math.pi * crack_depth_m)
            paris_c = positive_number(loading, "paris_c")
            paris_m = positive_number(loading, "paris_m")
            critical_depth = 0.8 * wall_thickness
            remaining_depth = max(critical_depth - controlling_crack["depth_mm"], 0.0)
            growth_per_cycle = paris_c * (delta_k**paris_m)
            remaining_cycles = remaining_depth / growth_per_cycle if growth_per_cycle > 0 else None

    max_equivalent_strain = 0.002 + 0.08 * max(a["depth_ratio"] for a in anomalies) + 0.035 * proximity + 0.01 * secondary_stress / max(smys, 1e-9)
    model_length = model_length_factor * diameter
    remote_mesh = max(diameter / 18.0, wall_thickness * 4.0)
    local_mesh = max(wall_thickness / 3.0, min(a["length_mm"] for a in anomalies) / 16.0, 1.0)
    remaining_ligament = min(wall_thickness - a["depth_mm"] for a in anomalies)
    through_ligament_elements = max(6, math.ceil(max(remaining_ligament, wall_thickness * 0.15) / max(local_mesh / 3.0, 0.25)))
    refinement = str(mesh.get("refinement", "standard"))
    refinement_multiplier = {"coarse": 0.75, "standard": 1.0, "fine": 1.35}.get(refinement, 1.0)
    estimated_elements = int(
        math.ceil(model_length / remote_mesh)
        * math.ceil(math.pi * diameter / remote_mesh)
        * through_ligament_elements
        * 0.45
        * refinement_multiplier
    )

    boundary_conditions = [
        "Left reference ring constrains rigid-body motion while avoiding local over-constraint.",
        "Internal pressure is applied to pipe ID and internal flaw surfaces.",
        "Closed-end pressure axial stress sigma = P*D/(4t) is included unless open-ended testing is selected.",
        "Remote pipe uses coarse mesh with smooth transition into a locally refined anomaly region.",
    ]
    if any(a["surface"] == "internal" and a["type"] in {"crack", "mixed"} for a in anomalies):
        boundary_conditions.append("Internal crack-face pressurization should be included where physically credible.")
    if residual_fraction > 0:
        boundary_conditions.append("Residual stress is included as a fraction of SMYS for weld, dent, repair, or cold-work influence.")

    return {
        "inputs": {
            "outside_diameter_mm": diameter,
            "wall_thickness_mm": wall_thickness,
            "maop_mpa": maop,
            "smys_mpa": smys,
            "smts_mpa": smts,
            "elastic_modulus_mpa": elastic_modulus,
            "fracture_toughness_mpa_sqrt_m": toughness,
            "model_length_factor": model_length_factor,
            "sizing_case": sizing_case,
        },
        "anomalies": anomalies,
        "outputs": {
            "status": status,
            "interaction_factor": interaction_factor,
            "interaction_classification": interaction_classification,
            "interaction_response": interaction_response,
            "weakest_isolated_failure_pressure_mpa": weakest_isolated,
            "combined_failure_pressure_mpa": combined_failure_pressure,
            "safety_factor": safety_factor,
            "safety_category": safety_category,
            "recommended_response": recommended_response,
            "axial_edge_spacing_mm": axial_spacing,
            "circumferential_edge_spacing_mm": circum_spacing,
            "lambda_x": lambda_x,
            "lambda_theta": lambda_theta,
            "proximity_index": proximity,
            "max_equivalent_plastic_strain": max_equivalent_strain,
            "k_max_mpa_sqrt_m": k_max,
            "remaining_cycles": remaining_cycles,
            "governing_failure_mode": "Fracture / crack driving force" if k_max > 0 and k_max > 0.7 * toughness else "Plastic collapse / local ligament strain",
            "critical_location": min(anomalies, key=lambda item: item["failure_pressure_mpa"])["id"],
        },
        "mesh": {
            "model_length_mm": model_length,
            "remote_element_size_mm": remote_mesh,
            "local_element_size_mm": local_mesh,
            "through_ligament_elements": through_ligament_elements,
            "estimated_solid_elements": estimated_elements,
            "refinement": refinement,
            "solver_strategy": str(mesh.get("solver_strategy", "implicit_riks")),
            "boundary_conditions": boundary_conditions,
        },
    }


ILI_SURROGATE_MODELS = {
    "ann": {
        "label": "ANN (4-8-1)",
        "weights": [
            [
                [-1.04943817, -0.24428267, 0.48011117, 0.64114515, -1.01045526, 0.46088008, -0.10581577, 0.00304953],
                [-0.61847442, -0.18924206, -0.10408139, -0.05462232, -0.0092246, 0.09579294, 0.06204165, 0.70625998],
                [-0.2062265, 1.68241021, 0.47480453, 0.21606125, -0.0077569, 0.2501094, 1.40753207, -0.36591954],
                [-0.72368866, -3.89640981, 0.29230347, -3.6002696, 0.66151664, -0.87387276, 1.09655055, 2.6131858],
            ],
            [[-0.51296449], [0.534795], [0.38438072], [0.40613434], [0.54463685], [-1.10627148], [-0.28277326], [-0.56608738]],
        ],
        "biases": [
            [0.00893461, 1.06319457, 0.30399444, 0.74088522, 0.01553316, -0.00965265, -0.24272274, 1.22800916],
            [0.25115384],
        ],
    },
    "dnn": {
        "label": "DNN (4-10-6-1)",
        "weights": [
            [
                [0.32970049, -0.80662362, 0.51312619, 0.31035763, -0.98018254, 0.18958865, -0.00201048, 0.84664847, -1.10094598, -0.58557949],
                [0.09701721, 0.36824157, -0.12536037, 0.22654306, -0.99076918, 2.32901163, -0.47888067, 0.10065089, 0.06049155, -0.18840999],
                [-0.56357619, 0.07798374, 0.87178049, 0.08183443, -0.10973698, -1.27009378, -0.03589889, -0.06085377, 0.06445468, -0.65780922],
                [0.28615494, -0.59955813, -1.58451849, -0.33541917, 1.46005577, 2.52034437, 0.73869995, -0.29410619, 0.02133744, -2.46967103],
            ],
            [
                [0.1559474, -0.21057451, -0.3597962, -0.06197005, -0.22106275, -0.27808019],
                [0.08659315, -0.42496249, 0.9041187, -0.53652431, -0.27400557, 0.32441114],
                [0.00839931, 0.07087696, 1.00809949, 0.08166608, -1.0161164, 0.78260791],
                [0.45123238, 0.24621551, 0.46702212, -0.06261163, 0.07435449, 0.09588483],
                [-0.96433326, -0.19520879, -0.23228061, 0.93036236, -0.17756866, -0.37243171],
                [1.63996979, -0.4625946, -1.23252162, 0.95722615, -0.19644295, 0.43787136],
                [-1.07818894, 0.06596407, -0.65336341, -0.00690185, -0.2219167, -0.0127365],
                [0.84275558, 0.52066923, -0.49216225, 0.29269691, -0.26475814, -0.17293531],
                [-0.41233179, -1.21723382, -0.20548696, -0.33399594, 0.01731332, 0.0778439],
                [-0.29324661, 1.03978757, 0.13387757, 0.85732968, 0.45768391, 0.19385297],
            ],
            [[-0.15303085], [-0.10181953], [0.38096908], [0.48826838], [-0.0069632], [0.312468]],
        ],
        "biases": [
            [-0.06587824, 0.14073653, 0.07710316, 0.05737755, -0.58514372, -0.52270639, -0.14987182, -0.11472667, 0.0129701, 0.33475232],
            [-0.2233051, -0.08298297, 0.48887013, 0.18079207, 0.0757206, 0.22179824],
            [0.29403004],
        ],
    },
}


ILI_SURROGATE_BENCHMARKS = [
    (0.2, 0.8, 3.0, 0.5, 0.74), (0.4, 0.8, 3.0, 0.5, 0.64), (0.5, 0.8, 3.0, 0.5, 0.56),
    (0.6, 0.8, 3.0, 0.5, 0.46), (0.8, 0.8, 3.0, 0.5, 0.26), (0.5, 0.2, 3.0, 0.5, 0.69),
    (0.5, 0.4, 3.0, 0.5, 0.62), (0.5, 1.2, 3.0, 0.5, 0.53), (0.5, 1.8, 3.0, 0.5, 0.52),
    (0.5, 0.8, 3.0, 0.2, 0.58), (0.5, 0.8, 3.0, 0.4, 0.57), (0.5, 0.8, 3.0, 0.6, 0.53),
    (0.5, 0.8, 3.0, 0.7, 0.46), (0.5, 0.8, 3.0, 0.8, 0.37), (0.5, 0.8, 3.0, 0.9, 0.20),
    (0.5, 0.8, 3.0, 1.0, 0.17), (0.2, 0.8, 1.0, 0.5, 0.75), (0.4, 0.8, 1.0, 0.5, 0.64),
    (0.5, 0.8, 1.0, 0.5, 0.56), (0.6, 0.8, 1.0, 0.5, 0.46), (0.8, 0.8, 1.0, 0.5, 0.27),
    (0.5, 0.2, 1.0, 0.5, 0.69), (0.5, 0.4, 1.0, 0.5, 0.61), (0.5, 1.2, 1.0, 0.5, 0.53),
    (0.5, 1.8, 1.0, 0.5, 0.52), (0.5, 0.8, 0.0, 0.0, 0.48), (0.5, 0.8, 0.5, 0.0, 0.50),
    (0.5, 0.8, 1.0, 0.0, 0.52), (0.5, 0.8, 2.0, 0.0, 0.53),
]


def ili_surrogate_predict(model_name: str, depth_ratio: float, length_ratio: float, spacing_ratio: float, axial_stress_ratio: float) -> float:
    model = ILI_SURROGATE_MODELS.get(model_name, ILI_SURROGATE_MODELS["ann"])
    values = [
        2.0 * (depth_ratio - 0.2) / 0.6 - 1.0,
        2.0 * (length_ratio - 0.2) / 1.6 - 1.0,
        2.0 * spacing_ratio / 3.0 - 1.0,
        2.0 * axial_stress_ratio - 1.0,
    ]
    for layer_index, (weights, biases) in enumerate(zip(model["weights"], model["biases"])):
        values = [
            sum(values[input_index] * weights[input_index][output_index] for input_index in range(len(values))) + biases[output_index]
            for output_index in range(len(biases))
        ]
        if layer_index < len(model["weights"]) - 1:
            values = [math.tanh(value) for value in values]
    return max(0.05, min(values[0], 1.05))


def ili_surrogate_validation(model_name: str) -> dict[str, Any]:
    actual = [item[4] for item in ILI_SURROGATE_BENCHMARKS]
    predicted = [ili_surrogate_predict(model_name, *item[:4]) for item in ILI_SURROGATE_BENCHMARKS]
    errors = [prediction - target for prediction, target in zip(predicted, actual)]
    mean_actual = sum(actual) / len(actual)
    ss_res = sum(error**2 for error in errors)
    ss_tot = sum((target - mean_actual) ** 2 for target in actual)
    return {
        "benchmark_case_count": len(actual),
        "r_squared": 1.0 - ss_res / ss_tot if ss_tot else 1.0,
        "mae_normalized_pressure": sum(abs(error) for error in errors) / len(errors),
        "mape_percent": sum(abs(error / target) for error, target in zip(errors, actual)) / len(errors) * 100.0,
        "maximum_absolute_error_percent": max(abs(error / target) for error, target in zip(errors, actual)) * 100.0,
        "benchmark_scope": "Reproduction check against published Table 4 FEA trends; not an independent certification dataset.",
        "published_ann_unseen_r_squared": 0.9921,
        "published_ann_unseen_error_range_percent": [-9.39, 4.63],
        "published_fea_burst_validation_max_difference_percent": 3.67,
        "source": "Lo, Karuppanan, and Ovinis (2021), doi:10.3390/jmse9030281",
    }


def inverse_standard_normal(probability: float) -> float:
    probability = min(max(probability, 1e-12), 1.0 - 1e-12)
    a = [-39.69683028665376, 220.9460984245205, -275.9285104469687, 138.357751867269, -30.66479806614716, 2.506628277459239]
    b = [-54.47609879822406, 161.5858368580409, -155.6989798598866, 66.80131188771972, -13.28068155288572]
    c = [-0.007784894002430293, -0.3223964580411365, -2.400758277161838, -2.549732539343734, 4.374664141464968, 2.938163982698783]
    d = [0.007784695709041462, 0.3224671290700398, 2.445134137142996, 3.754408661907416]
    low = 0.02425
    high = 1.0 - low
    if probability < low:
        q = math.sqrt(-2.0 * math.log(probability))
        return (((((c[0] * q + c[1]) * q + c[2]) * q + c[3]) * q + c[4]) * q + c[5]) / ((((d[0] * q + d[1]) * q + d[2]) * q + d[3]) * q + 1)
    if probability <= high:
        q = probability - 0.5
        r = q * q
        return (((((a[0] * r + a[1]) * r + a[2]) * r + a[3]) * r + a[4]) * r + a[5]) * q / (((((b[0] * r + b[1]) * r + b[2]) * r + b[3]) * r + b[4]) * r + 1)
    q = math.sqrt(-2.0 * math.log(1.0 - probability))
    return -(((((c[0] * q + c[1]) * q + c[2]) * q + c[3]) * q + c[4]) * q + c[5]) / ((((d[0] * q + d[1]) * q + d[2]) * q + d[3]) * q + 1)


def calculate_ili_to_fea_payload(payload: dict[str, Any]) -> dict[str, Any]:
    pipe = payload.get("pipe", {})
    loading = payload.get("loading", {})
    model = payload.get("model", {})
    features = payload.get("features", {})
    raw_data = payload.get("raw_data", {})
    weld = payload.get("weld", {})

    diameter = positive_number(pipe, "outside_diameter_mm")
    wall_thickness = positive_number(pipe, "wall_thickness_mm")
    maop = positive_number(pipe, "maop_mpa")
    smys = positive_number(pipe, "smys_mpa")
    smts = positive_number(pipe, "smts_mpa")
    elastic_modulus = positive_number(pipe, "elastic_modulus_mpa")
    toughness = positive_number(pipe, "fracture_toughness_mpa_sqrt_m")
    assessment_factor = float(pipe.get("assessment_factor", 0.72) or 0.72)
    if assessment_factor <= 0 or assessment_factor > 1:
        raise ValueError("assessment_factor must be greater than zero and no more than one")
    weld_yield = float(weld.get("yield_strength_mpa", smys) or smys)
    weld_toughness = float(weld.get("fracture_toughness_mpa_sqrt_m", toughness) or toughness)
    weld_residual_factor = float(weld.get("residual_stress_factor", 0.6) or 0.0)
    weld_cap_width = float(weld.get("cap_width_mm", 12.0) or 12.0)
    haz_width = float(weld.get("haz_width_mm", 6.0) or 0.0)
    pipe_weld_type = str(weld.get("pipe_weld_type", "")).strip().lower()
    legacy_seam_type = str(weld.get("seam_type", "longitudinal_seam")).strip().lower()
    legacy_process = str(weld.get("manufacturing_process", "erw")).strip().lower()
    if not pipe_weld_type:
        pipe_weld_type = (
            "seamless"
            if legacy_seam_type == "none"
            else "spiral_dsaw"
            if legacy_seam_type == "spiral_seam"
            else "long_seam_dsaw"
            if legacy_process == "dsaw"
            else "erw_high_frequency"
        )
    weld_type_map = {
        "erw_high_frequency": ("longitudinal_seam", "erw_high_frequency", "erw"),
        "erw_low_frequency": ("longitudinal_seam", "erw_low_frequency", "erw"),
        "long_seam_dsaw": ("longitudinal_seam", "long_seam_dsaw", "dsaw"),
        "spiral_dsaw": ("spiral_seam", "spiral_dsaw", "dsaw"),
        "electric_flash_welded": ("longitudinal_seam", "electric_flash_welded", "legacy_flush"),
        "furnace_butt_welded": ("longitudinal_seam", "furnace_butt_welded", "legacy_flush"),
        "seamless": ("none", "seamless", "seamless"),
    }
    if pipe_weld_type not in weld_type_map:
        raise ValueError("pipe_weld_type is not supported")
    pipe_seam_type, manufacturing_process, manufacturing_family = weld_type_map[pipe_weld_type]
    spiral_turns = float(weld.get("spiral_turns", 1.5) or 1.5)
    if spiral_turns < 0.25 or spiral_turns > 10:
        raise ValueError("spiral_turns must be between 0.25 and 10")
    if weld_yield <= 0 or weld_toughness <= 0:
        raise ValueError("Weld yield strength and fracture toughness must be greater than zero")
    if weld_residual_factor < 0 or weld_residual_factor > 1:
        raise ValueError("Weld residual stress factor must be between zero and one")
    if weld_cap_width <= 0 or haz_width < 0:
        raise ValueError("Weld cap width must be greater than zero and HAZ width cannot be negative")

    ids = [str(value).strip() for value in features.get("ids", [])]
    types = [str(value).strip().lower() for value in features.get("types", [])]
    depths_percent = [float(value) for value in features.get("depths_percent", [])]
    lengths_mm = [float(value) for value in features.get("lengths_mm", [])]
    distances_m = [float(value) for value in features.get("distances_m", [])]
    clock_positions = [str(value).strip() for value in features.get("clock_positions", [])]
    widths_mm = [float(value) for value in features.get("widths_mm", [])]
    orientations_deg = [float(value) for value in features.get("orientations_deg", [])]
    surfaces = [str(value).strip().lower() for value in features.get("surfaces", [])]
    weld_types = [str(value).strip().lower() for value in features.get("weld_types", [])]
    weld_offsets_mm = [float(value) for value in features.get("weld_offsets_mm", [])]
    reported_pressures = [float(value) if value not in (None, "") else 0.0 for value in features.get("reported_failure_pressures_mpa", [])]
    geometry_source = str(model.get("geometry_source", "auto")).strip().lower()
    if geometry_source not in {"auto", "raw", "feature"}:
        raise ValueError("geometry_source must be auto, raw, or feature")

    raw_groups: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for source_name, items in (
        ("mfl", raw_data.get("mfl_samples", []) or []),
        ("crack", raw_data.get("crack_samples", []) or []),
        ("caliper", raw_data.get("caliper_samples", []) or []),
    ):
        for raw_index, raw in enumerate(items):
            feature_id = str(raw.get("feature_id", "")).strip() or f"{source_name.upper()}-001"
            raw["feature_id"] = feature_id
            raw_groups.setdefault(feature_id, {"mfl": [], "crack": [], "caliper": []})[source_name].append(raw)

    if geometry_source == "raw" or not ids:
        if not raw_groups:
            raise ValueError("Raw tool data is required when the feature list is omitted or Raw tool data only is selected")
        ids = list(raw_groups)
        types = []
        depths_percent = []
        lengths_mm = []
        distances_m = []
        clock_positions = []
        widths_mm = []
        orientations_deg = []
        surfaces = []
        weld_types = []
        weld_offsets_mm = []
        reported_pressures = []
        for feature_id in ids:
            grouped = raw_groups[feature_id]
            all_samples = grouped["mfl"] + grouped["crack"] + grouped["caliper"]
            crack_ids = {str(item.get("crack_id", "")).strip() for item in grouped["crack"] if str(item.get("crack_id", "")).strip()}
            crack_types = {str(item.get("anomaly_type", "")).strip().lower() for item in grouped["crack"]}
            kind = (
                "scc"
                if "scc" in crack_types or len(crack_ids) > 1
                else "crack"
                if grouped["crack"]
                else "metal_loss"
                if grouped["mfl"]
                else "dent"
            )
            axial_values = [float(item.get("axial_offset_mm", 0.0) or 0.0) for item in all_samples]
            circumferential_values = [float(item.get("circumferential_offset_mm", 0.0) or 0.0) for item in all_samples]
            axial_span = max(axial_values) - min(axial_values) if len(axial_values) > 1 else 0.0
            circumferential_span = (
                max(circumferential_values) - min(circumferential_values) if len(circumferential_values) > 1 else 0.0
            )
            if grouped["mfl"]:
                depth_percent = max(float(item.get("depth_percent", 0.0) or 0.0) for item in grouped["mfl"])
            elif grouped["crack"]:
                depth_percent = max(float(item.get("depth_mm", 0.0) or 0.0) for item in grouped["crack"]) / wall_thickness * 100.0
            else:
                depth_percent = (
                    max(abs(float(item.get("radial_deformation_mm", 0.0) or 0.0)) for item in grouped["caliper"])
                    / wall_thickness
                    * 100.0
                )
            distance = next((float(item.get("distance_m", 0.0) or 0.0) for item in all_samples if float(item.get("distance_m", 0.0) or 0.0)), 0.0)
            clock = next((str(item.get("clock_position", "")).strip() for item in all_samples if str(item.get("clock_position", "")).strip()), "12:00")
            orientation = next(
                (float(item.get("orientation_deg", 0.0) or 0.0) for item in grouped["crack"] if float(item.get("orientation_deg", 0.0) or 0.0)),
                0.0,
            )
            types.append(kind)
            depths_percent.append(min(max(depth_percent, 0.01), 99.0))
            lengths_mm.append(max(axial_span, wall_thickness * (3.0 if kind in {"crack", "scc"} else 4.0)))
            widths_mm.append(max(circumferential_span, wall_thickness * (1.5 if kind in {"crack", "scc"} else 3.0)))
            distances_m.append(distance)
            clock_positions.append(clock)
            orientations_deg.append(orientation)
            surfaces.append("external")
            weld_types.append("none")
            weld_offsets_mm.append(0.0)
            reported_pressures.append(0.0)

    count = len(ids)
    if count == 0:
        raise ValueError("At least one ILI feature is required")
    required = {
        "types": types,
        "depths_percent": depths_percent,
        "lengths_mm": lengths_mm,
        "distances_m": distances_m,
        "clock_positions": clock_positions,
    }
    for name, values in required.items():
        if len(values) != count:
            raise ValueError(f"{name} must contain one value per ILI feature")
    if widths_mm and len(widths_mm) != count:
        raise ValueError("widths_mm must be empty or contain one value per ILI feature")
    if orientations_deg and len(orientations_deg) != count:
        raise ValueError("orientations_deg must be empty or contain one value per ILI feature")
    if surfaces and len(surfaces) != count:
        raise ValueError("surfaces must be empty or contain one value per ILI feature")
    if weld_types and len(weld_types) != count:
        raise ValueError("weld_types must be empty or contain one value per ILI feature")
    if weld_offsets_mm and len(weld_offsets_mm) != count:
        raise ValueError("weld_offsets_mm must be empty or contain one value per ILI feature")
    if reported_pressures and len(reported_pressures) != count:
        raise ValueError("reported_failure_pressures_mpa must be empty or contain one value per ILI feature")

    widths_mm = widths_mm or [0.0] * count
    orientations_deg = orientations_deg or [0.0] * count
    surfaces = surfaces or ["external"] * count
    weld_types = weld_types or ["none"] * count
    weld_offsets_mm = weld_offsets_mm or [0.0] * count
    reported_pressures = reported_pressures or [0.0] * count

    def clock_to_degrees(value: str) -> float:
        text = value.strip()
        if ":" in text:
            hour_text, minute_text = text.split(":", 1)
            hour = float(hour_text or 0) % 12
            minute = float(minute_text or 0)
            return (hour * 30.0 + minute * 0.5) % 360.0
        return float(text or 0.0) % 360.0

    minimum_distance = min(distances_m)
    translated = []
    for index, feature_id in enumerate(ids):
        depth_percent = depths_percent[index]
        length = lengths_mm[index]
        if depth_percent <= 0 or depth_percent >= 100:
            raise ValueError("Each ILI depth percentage must be greater than zero and less than 100")
        if length <= 0:
            raise ValueError("Each ILI feature length must be greater than zero")
        kind = types[index]
        weld_type = weld_types[index].replace("-", "_").replace(" ", "_")
        if weld_type == "pipe_seam":
            weld_type = pipe_seam_type
        if weld_type not in {"none", "girth_weld", "longitudinal_seam", "spiral_seam"}:
            raise ValueError("Each weld type must be none, girth_weld, pipe_seam, longitudinal_seam, or spiral_seam")
        weld_offset = weld_offsets_mm[index]
        weld_zone = (
            "base_metal"
            if weld_type == "none" or abs(weld_offset) > weld_cap_width / 2.0 + haz_width
            else "weld_metal"
            if abs(weld_offset) <= weld_cap_width / 2.0
            else "haz"
        )
        fea_type = "crack" if kind in {"crack", "crack_like", "scc"} else kind if kind in {"metal_loss", "dent", "weld", "mixed"} else "mixed"
        inferred_width = max(
            wall_thickness * (2.0 if fea_type == "crack" else 4.0),
            length * (0.2 if fea_type == "crack" else 0.55),
        )
        translated.append(
            {
                "id": feature_id,
                "source_index": index,
                "type": fea_type,
                "source_type": kind,
                "surface": surfaces[index] if surfaces[index] in {"external", "internal"} else "external",
                "axial_location_mm": (distances_m[index] - minimum_distance) * 1000.0,
                "absolute_distance_m": distances_m[index],
                "clock_position_deg": clock_to_degrees(clock_positions[index]),
                "clock_position": clock_positions[index],
                "length_mm": length,
                "width_mm": widths_mm[index] if widths_mm[index] > 0 else inferred_width,
                "depth_mm": depth_percent * wall_thickness / 100.0,
                "depth_percent": depth_percent,
                "orientation_deg": orientations_deg[index],
                "weld_type": weld_type,
                "weld_offset_mm": weld_offset,
                "weld_zone": weld_zone,
                "manufacturing_process": "field_girth_weld" if weld_type == "girth_weld" else manufacturing_process if weld_type != "none" else "n/a",
                "manufacturing_family": "field_girth_weld" if weld_type == "girth_weld" else manufacturing_family if weld_type != "none" else "n/a",
                "spiral_turns": spiral_turns,
            }
        )

    translated_by_id = {feature["id"]: feature for feature in translated}

    def raw_samples(name: str, value_keys: tuple[str, ...]) -> list[dict[str, Any]]:
        samples = []
        for index, raw in enumerate(raw_data.get(name, []) or []):
            feature_id = str(raw.get("feature_id", "")).strip()
            if feature_id not in translated_by_id:
                raise ValueError(f"{name} sample {index + 1} references unknown feature {feature_id or '(blank)'}")
            sample = {
                "feature_id": feature_id,
                "axial_offset_mm": float(raw.get("axial_offset_mm", 0.0) or 0.0),
                "circumferential_offset_mm": float(raw.get("circumferential_offset_mm", 0.0) or 0.0),
                "distance_m": float(raw.get("distance_m", 0.0) or 0.0),
                "clock_position": str(raw.get("clock_position", "")).strip(),
                "orientation_deg": float(raw.get("orientation_deg", 0.0) or 0.0),
                "crack_id": str(raw.get("crack_id", "")).strip(),
                "anomaly_type": str(raw.get("anomaly_type", "")).strip().lower(),
            }
            for key in value_keys:
                sample[key] = float(raw.get(key, 0.0) or 0.0)
            samples.append(sample)
        return samples

    mfl_samples = raw_samples("mfl_samples", ("depth_percent",))
    crack_samples = raw_samples("crack_samples", ("depth_mm", "opening_mm"))
    caliper_samples = raw_samples("caliper_samples", ("radial_deformation_mm",))
    raw_sample_count = len(mfl_samples) + len(crack_samples) + len(caliper_samples)
    raw_by_feature: dict[str, dict[str, list[dict[str, Any]]]] = {
        feature_id: {"mfl": [], "crack": [], "caliper": []} for feature_id in ids
    }
    for sample in mfl_samples:
        raw_by_feature[sample["feature_id"]]["mfl"].append(sample)
    for sample in crack_samples:
        raw_by_feature[sample["feature_id"]]["crack"].append(sample)
    for sample in caliper_samples:
        raw_by_feature[sample["feature_id"]]["caliper"].append(sample)
    for feature in translated:
        feature["raw_sample_counts"] = {
            source: len(raw_by_feature[feature["id"]][source]) for source in ("mfl", "crack", "caliper")
        }
        grouped = raw_by_feature[feature["id"]]
        all_samples = grouped["mfl"] + grouped["crack"] + grouped["caliper"]
        if geometry_source != "feature" and all_samples:
            axial_values = [item["axial_offset_mm"] for item in all_samples]
            circumferential_values = [item["circumferential_offset_mm"] for item in all_samples]
            axial_span = max(axial_values) - min(axial_values) if len(axial_values) > 1 else 0.0
            circumferential_span = (
                max(circumferential_values) - min(circumferential_values) if len(circumferential_values) > 1 else 0.0
            )
            mfl_depths = [item["depth_percent"] for item in grouped["mfl"]]
            crack_depths = [item["depth_mm"] for item in grouped["crack"]]
            dent_depths = [abs(item["radial_deformation_mm"]) for item in grouped["caliper"]]
            raw_depth_mm = (
                max(mfl_depths) * wall_thickness / 100.0
                if mfl_depths
                else max(crack_depths)
                if crack_depths
                else max(dent_depths)
                if dent_depths
                else feature["depth_mm"]
            )
            feature["length_mm"] = max(feature["length_mm"], axial_span)
            feature["width_mm"] = max(feature["width_mm"], circumferential_span)
            feature["depth_mm"] = min(raw_depth_mm, wall_thickness * 0.99)
            feature["depth_percent"] = feature["depth_mm"] / wall_thickness * 100.0
            if grouped["crack"]:
                crack_ids = {item["crack_id"] for item in grouped["crack"] if item["crack_id"]}
                is_scc = any(item["anomaly_type"] == "scc" for item in grouped["crack"]) or len(crack_ids) > 1
                feature["type"] = "crack"
                feature["source_type"] = "scc" if is_scc else "crack"
            elif grouped["mfl"]:
                feature["type"] = "metal_loss"
                feature["source_type"] = "metal_loss"
            elif grouped["caliper"]:
                feature["type"] = "dent"
                feature["source_type"] = "dent"
            feature["geometry_source"] = "raw_tool_data"
        else:
            feature["geometry_source"] = "feature_list"
        depths = [item["depth_percent"] for item in grouped["mfl"]]
        feature["effective_geometry"] = {
            "minimum_remaining_ligament_mm": max(wall_thickness - feature["depth_mm"], wall_thickness * 0.01),
            "metal_loss_area_ratio": (
                sum(depths) / (len(depths) * 100.0) if depths else feature["depth_percent"] / 100.0
            ),
            "maximum_crack_opening_mm": max((item["opening_mm"] for item in grouped["crack"]), default=0.0),
            "crack_front_points": len(grouped["crack"]),
            "caliper_points": len(grouped["caliper"]),
            "mfl_points": len(grouped["mfl"]),
            "mesh_void_required": bool(grouped["crack"]),
        }

    types = [feature["source_type"] for feature in translated]
    depths_percent = [feature["depth_percent"] for feature in translated]
    lengths_mm = [feature["length_mm"] for feature in translated]
    widths_mm = [feature["width_mm"] for feature in translated]
    orientations_deg = [feature["orientation_deg"] for feature in translated]
    surfaces = [feature["surface"] for feature in translated]

    cycles_per_year = max(0.0, float(loading.get("cycles_per_year", 1000.0) or 1000.0))
    pressure_range = max(0.0, float(loading.get("pressure_range_mpa", 0.0) or 0.0))
    surrogate_name = str(model.get("surrogate_model", "ann")).strip().lower()
    if surrogate_name not in ILI_SURROGATE_MODELS:
        raise ValueError("surrogate_model must be ann or dnn")
    reliability_enabled = bool(model.get("reliability_enabled", True))
    reliability_samples = int(float(model.get("reliability_samples", 2500) or 2500))
    if reliability_enabled and (reliability_samples < 500 or reliability_samples > 50000):
        raise ValueError("reliability_samples must be between 500 and 50000")
    depth_cov = max(0.0, float(model.get("depth_cov", 0.10) or 0.0))
    pressure_cov = max(0.0, float(model.get("pressure_cov", 0.03) or 0.0))
    model_error_cov = max(0.0, float(model.get("model_error_cov", 0.08) or 0.0))
    strain_limit = float(model.get("strain_limit", 0.06) or 0.06)
    if strain_limit <= 0:
        raise ValueError("strain_limit must be greater than zero")
    screening = calculate_ili_screening_payload(
        {
            "pipe": {
                "outside_diameter_mm": diameter,
                "wall_thickness_mm": wall_thickness,
                "maop_mpa": maop,
                "smys_mpa": smys,
                "assessment_factor": assessment_factor,
            },
            "criteria": {
                "repair_pressure_ratio": 1.0,
                "monitor_pressure_ratio": 0.8,
                "depth_watch_percent": 50.0,
                "primary_method": str(model.get("screening_method", "modified_b31g")),
                "screening_methods": [
                    "modified_b31g",
                    "asme_b31g",
                    "rstreng_simplified",
                    "corlas",
                    "scc_colony",
                    "reported_pressure",
                    "crack_fracture",
                ],
                "fracture_toughness_mpa_sqrt_m": toughness,
            },
            "fatigue": {
                "enabled": pressure_range > 0,
                "stress_range_mpa": max(pressure_range * diameter / (2.0 * wall_thickness), 0.0),
                "bending_strain_percent": float(loading.get("bending_strain_percent", 0.0) or 0.0),
                "cycles_per_year": cycles_per_year,
                "applied_cycles": float(loading.get("applied_cycles", 0.0) or 0.0),
                "paris_c": float(loading.get("paris_c", 1e-12) or 1e-12),
                "paris_m": float(loading.get("paris_m", 3.0) or 3.0),
            },
            "risk": {
                "class_location": str(model.get("class_location", "1")),
                "prediction_years": float(model.get("prediction_years", 5.0) or 5.0),
                "annual_growth_percent": float(model.get("annual_growth_percent", 0.0) or 0.0),
            },
            "features": {
                "ids": ids,
                "types": types,
                "depths_percent": depths_percent,
                "lengths_mm": lengths_mm,
                "clock_positions": clock_positions,
                "distances_m": distances_m,
                "reported_failure_pressures_mpa": reported_pressures,
            },
        }
    )
    screening_by_id = {item["feature_id"]: item for item in screening["outputs"]["ranked_features"]}

    def weld_crack_failure_pressure(feature: dict[str, Any]) -> tuple[float, float, float]:
        crack_depth_m = max(feature["depth_mm"], 0.01) / 1000.0
        residual_stress = weld_yield * weld_residual_factor

        def fad_margin(pressure: float) -> tuple[float, float, float]:
            hoop_stress = pressure * diameter / (2.0 * wall_thickness)
            axial_stress = pressure * diameter / (4.0 * wall_thickness)
            if feature["weld_type"] == "girth_weld":
                primary_stress = axial_stress
            elif feature["weld_type"] == "spiral_seam":
                angle = math.radians(45.0)
                primary_stress = math.sqrt((hoop_stress * math.cos(angle)) ** 2 + (axial_stress * math.sin(angle)) ** 2)
            else:
                primary_stress = hoop_stress
            lr = primary_stress / max(weld_yield, 1e-9)
            kr = (
                1.12
                * (primary_stress + residual_stress)
                * math.sqrt(math.pi * crack_depth_m)
                / max(weld_toughness, 1e-9)
            )
            boundary = max(0.0, (1.0 - 0.14 * lr**2) * (0.3 + 0.7 * math.exp(-0.65 * lr**6)))
            return boundary - kr, lr, kr

        lower = 0.0
        upper = max(pristine_pressure if "pristine_pressure" in locals() else 2.0 * wall_thickness * smts / diameter, maop * 3.0)
        for _ in range(70):
            midpoint = (lower + upper) / 2.0
            margin, lr, _ = fad_margin(midpoint)
            if margin >= 0 and lr <= 1.2:
                lower = midpoint
            else:
                upper = midpoint
        _, final_lr, final_kr = fad_margin(lower)
        return lower, final_lr, final_kr

    weld_assessments = []
    weld_assessment_by_id: dict[str, dict[str, Any]] = {}
    for feature in translated:
        if feature["weld_zone"] == "base_metal":
            continue
        generic = screening_by_id[feature["id"]]["governing_failure_pressure_mpa"]
        process_label = feature["manufacturing_process"].replace("_", " ").upper()
        if feature["type"] == "crack":
            failure_pressure, load_ratio, fracture_ratio = weld_crack_failure_pressure(feature)
            method = f"{process_label} FAD - {feature['weld_type']}"
            mode = "Weld/HAZ fracture"
        else:
            zone_factor = 0.88 if feature["weld_zone"] == "weld_metal" else 0.93
            mismatch_factor = min(max(weld_yield / smys, 0.65), 1.05)
            residual_capacity_factor = 1.0 / (1.0 + 0.25 * weld_residual_factor)
            failure_pressure = generic * zone_factor * mismatch_factor * residual_capacity_factor
            load_ratio = failure_pressure / max(2.0 * wall_thickness * weld_yield / diameter, 1e-9)
            fracture_ratio = None
            method = f"{process_label} modified capacity - {feature['weld_type']}"
            mode = "Weld/HAZ local collapse"
        assessment = {
            "feature_id": feature["id"],
            "weld_type": feature["weld_type"],
            "weld_zone": feature["weld_zone"],
            "weld_offset_mm": feature["weld_offset_mm"],
            "manufacturing_process": feature["manufacturing_process"],
            "calculation_method": method,
            "failure_mode": mode,
            "failure_pressure_mpa": failure_pressure,
            "maximum_mop_mpa": failure_pressure * assessment_factor,
            "load_ratio": load_ratio,
            "fracture_ratio": fracture_ratio,
            "weld_strength_mismatch": weld_yield / smys,
            "residual_stress_mpa": weld_yield * weld_residual_factor,
            "basis": (
                "Crack-like weld defects use a Level-2-style failure-assessment diagram with pressure stress, "
                "weld residual stress, and weld toughness."
                if feature["type"] == "crack"
                else "Weld metal and HAZ defects use the isolated anomaly pressure with weld-zone, strength-mismatch, "
                "and residual-stress capacity modifiers."
            ),
        }
        weld_assessments.append(assessment)
        weld_assessment_by_id[feature["id"]] = assessment

    interaction_distance = max(
        0.0,
        float(model.get("interaction_distance_mm", 3.0 * math.sqrt(diameter * wall_thickness)) or 0.0),
    )
    pair_models = []
    for left_index, left in enumerate(translated):
        for right in translated[left_index + 1 :]:
            center_distance = abs(right["axial_location_mm"] - left["axial_location_mm"])
            axial_edge_spacing = max(0.0, center_distance - (left["length_mm"] + right["length_mm"]) / 2.0)
            if axial_edge_spacing > interaction_distance:
                continue
            interaction = calculate_interacting_anomalies_payload(
                {
                    "pipe": {
                        "outside_diameter_mm": diameter,
                        "wall_thickness_mm": wall_thickness,
                        "maop_mpa": maop,
                        "smys_mpa": smys,
                        "smts_mpa": smts,
                        "elastic_modulus_mpa": elastic_modulus,
                        "fracture_toughness_mpa_sqrt_m": toughness,
                        "model_length_factor": float(model.get("model_length_factor", 8.0) or 8.0),
                    },
                    "loading": {
                        "secondary_stress_mpa": float(loading.get("secondary_stress_mpa", 0.0) or 0.0),
                        "residual_stress_fraction": float(loading.get("residual_stress_fraction", 0.0) or 0.0),
                        "pressure_range_mpa": pressure_range,
                        "paris_c": float(loading.get("paris_c", 1e-12) or 1e-12),
                        "paris_m": float(loading.get("paris_m", 3.0) or 3.0),
                    },
                    "uncertainty": {
                        "case": str(model.get("sizing_case", "conservative")),
                        "depth_tolerance_mm": float(model.get("depth_tolerance_mm", 0.0) or 0.0),
                        "length_tolerance_mm": float(model.get("length_tolerance_mm", 0.0) or 0.0),
                        "width_tolerance_mm": float(model.get("width_tolerance_mm", 0.0) or 0.0),
                    },
                    "mesh": {
                        "refinement": str(model.get("mesh_refinement", "standard")),
                        "solver_strategy": str(model.get("solver_strategy", "implicit_riks")),
                    },
                    "anomalies": [left, right],
                }
            )
            pair_output = interaction["outputs"]
            pair_failure_pressure = pair_output["combined_failure_pressure_mpa"]
            weld_pair_factors = []
            for pair_feature in (left, right):
                weld_result = weld_assessment_by_id.get(pair_feature["id"])
                if weld_result:
                    generic_pressure = screening_by_id[pair_feature["id"]]["governing_failure_pressure_mpa"]
                    weld_pair_factors.append(weld_result["failure_pressure_mpa"] / max(generic_pressure, 1e-9))
            weld_interaction_factor = min(weld_pair_factors, default=1.0)
            pair_failure_pressure *= weld_interaction_factor
            allowable = pair_failure_pressure * assessment_factor
            remaining_cycles = pair_output["remaining_cycles"]
            pair_models.append(
                {
                    "model_id": f"FEA-{left['id']}-{right['id']}",
                    "feature_ids": [left["id"], right["id"]],
                    "interaction_classification": pair_output["interaction_classification"],
                    "interaction_factor": pair_output["interaction_factor"],
                    "combined_failure_pressure_mpa": pair_failure_pressure,
                    "maximum_mop_mpa": allowable,
                    "maop_ratio": maop / allowable if allowable > 0 else float("inf"),
                    "remaining_cycles": remaining_cycles,
                    "fatigue_life_years": remaining_cycles / cycles_per_year if remaining_cycles is not None and cycles_per_year > 0 else None,
                    "critical_location": pair_output["critical_location"],
                    "governing_failure_mode": (
                        f"{pair_output['governing_failure_mode']} with weld/HAZ modifier"
                        if weld_pair_factors
                        else pair_output["governing_failure_mode"]
                    ),
                    "weld_interaction_factor": weld_interaction_factor,
                    "maximum_equivalent_plastic_strain": pair_output["max_equivalent_plastic_strain"],
                    "mesh": interaction["mesh"],
                    "anomalies": interaction["anomalies"],
                }
            )

    sqrt_dt = math.sqrt(diameter * wall_thickness)
    pristine_pressure = 2.0 * wall_thickness * smts / diameter
    axial_stress_ratio = min(
        abs(float(loading.get("secondary_stress_mpa", 0.0) or 0.0)) / max(smys, 1e-9),
        1.0,
    )
    surrogate_predictions = []
    for feature in translated:
        if feature["type"] != "metal_loss" or feature["weld_zone"] != "base_metal":
            continue
        other_features = [item for item in translated if item["id"] != feature["id"]]
        spacing_ratio = 3.0
        if other_features:
            nearest = min(
                max(
                    0.0,
                    abs(item["axial_location_mm"] - feature["axial_location_mm"])
                    - (item["length_mm"] + feature["length_mm"]) / 2.0,
                )
                for item in other_features
            )
            spacing_ratio = min(nearest / max(sqrt_dt, 1e-9), 3.0)
        depth_ratio = feature["depth_mm"] / wall_thickness
        length_ratio = feature["length_mm"] / diameter
        within_domain = 0.2 <= depth_ratio <= 0.8 and 0.2 <= length_ratio <= 1.8 and 0 <= spacing_ratio <= 3 and axial_stress_ratio <= 1
        normalized_pressure = ili_surrogate_predict(
            surrogate_name,
            min(max(depth_ratio, 0.2), 0.8),
            min(max(length_ratio, 0.2), 1.8),
            min(max(spacing_ratio, 0.0), 3.0),
            axial_stress_ratio,
        )
        predicted_pressure = normalized_pressure * pristine_pressure
        lower_confidence_pressure = predicted_pressure * max(0.5, 1.0 - 1.645 * model_error_cov)
        surrogate_predictions.append(
            {
                "feature_id": feature["id"],
                "model": ILI_SURROGATE_MODELS[surrogate_name]["label"],
                "normalized_failure_pressure": normalized_pressure,
                "predicted_failure_pressure_mpa": predicted_pressure,
                "lower_95_failure_pressure_mpa": lower_confidence_pressure,
                "advisory_maximum_mop_mpa": lower_confidence_pressure * assessment_factor,
                "within_training_domain": within_domain,
                "inputs": {
                    "depth_to_thickness": depth_ratio,
                    "length_to_diameter": length_ratio,
                    "spacing_to_sqrt_dt": spacing_ratio,
                    "axial_stress_to_smys": axial_stress_ratio,
                },
            }
        )

    reliability = {
        "available": False,
        "enabled": reliability_enabled,
        "probability_of_failure": None,
        "reliability_index_beta": None,
        "samples": reliability_samples if reliability_enabled else 0,
        "basis": (
            "Reliability module was deactivated by the user."
            if not reliability_enabled
            else "No metal-loss feature was available for the published corrosion surrogate domain."
        ),
    }
    if reliability_enabled and surrogate_predictions:
        controlling_surrogate = min(surrogate_predictions, key=lambda item: item["lower_95_failure_pressure_mpa"])
        feature = translated_by_id[controlling_surrogate["feature_id"]]
        inputs = controlling_surrogate["inputs"]
        rng = random.Random(int(float(model.get("reliability_seed", 1183) or 1183)))
        failure_count = 0
        capacities = []
        for _ in range(reliability_samples):
            sampled_depth = min(max(rng.gauss(inputs["depth_to_thickness"], max(inputs["depth_to_thickness"] * depth_cov, 1e-6)), 0.2), 0.8)
            sampled_length = min(max(rng.gauss(inputs["length_to_diameter"], max(inputs["length_to_diameter"] * depth_cov * 0.6, 1e-6)), 0.2), 1.8)
            normalized = ili_surrogate_predict(
                surrogate_name,
                sampled_depth,
                sampled_length,
                inputs["spacing_to_sqrt_dt"],
                inputs["axial_stress_to_smys"],
            )
            capacity = normalized * pristine_pressure * max(rng.gauss(1.0, model_error_cov), 0.45)
            demand = maop * max(rng.gauss(1.0, pressure_cov), 0.7)
            capacities.append(capacity)
            if demand >= capacity:
                failure_count += 1
        probability_failure = failure_count / reliability_samples
        plotting_probability = (failure_count + 0.5) / (reliability_samples + 1.0)
        capacities.sort()
        reliability = {
            "available": True,
            "enabled": True,
            "controlling_feature": feature["id"],
            "probability_of_failure": probability_failure,
            "reliability_index_beta": -inverse_standard_normal(plotting_probability),
            "samples": reliability_samples,
            "p05_capacity_mpa": capacities[max(0, math.ceil(0.05 * reliability_samples) - 1)],
            "median_capacity_mpa": capacities[reliability_samples // 2],
            "depth_cov": depth_cov,
            "pressure_cov": pressure_cov,
            "model_error_cov": model_error_cov,
            "basis": "Monte Carlo propagation through the selected corrosion surrogate; advisory and domain-limited.",
        }

    dent_strains = []
    for feature in translated:
        if feature["type"] != "dent":
            continue
        half_length = max(feature["length_mm"] / 2.0, wall_thickness)
        half_width = max(feature["width_mm"] / 2.0, wall_thickness)
        curvature = 2.0 * feature["depth_mm"] / half_length**2 + 2.0 * feature["depth_mm"] / half_width**2
        dent_strains.append(
            {
                "feature_id": feature["id"],
                "geometric_bending_strain": wall_thickness * curvature / 2.0,
            }
        )
    interaction_strain = max(
        (item["maximum_equivalent_plastic_strain"] for item in pair_models),
        default=0.0,
    )
    existing_bending_strain = max(0.0, float(loading.get("bending_strain_percent", 0.0) or 0.0)) / 100.0
    maximum_dent_strain = max((item["geometric_bending_strain"] for item in dent_strains), default=0.0)
    maximum_strain = max(existing_bending_strain + maximum_dent_strain, interaction_strain)
    has_dent_coincident_feature = any(
        {"dent", "crack"}.issubset({anomaly["type"] for anomaly in item["anomalies"]})
        or {"dent", "metal_loss"}.issubset({anomaly["type"] for anomaly in item["anomalies"]})
        for item in pair_models
    )
    b318_status = "ACCEPTABLE" if maximum_strain <= strain_limit else "ACTION REQUIRED"
    api_1183_status = (
        "DETAILED ASSESSMENT REQUIRED"
        if has_dent_coincident_feature or maximum_strain > strain_limit
        else "SCREENING ACCEPTABLE - FATIGUE REVIEW REQUIRED"
        if dent_strains
        else "NOT APPLICABLE"
    )
    strain_assessment = {
        "maximum_strain": maximum_strain,
        "maximum_strain_percent": maximum_strain * 100.0,
        "strain_limit": strain_limit,
        "strain_limit_percent": strain_limit * 100.0,
        "utilization": maximum_strain / strain_limit,
        "b31_8_status": b318_status,
        "api_rp_1183_status": api_1183_status,
        "dent_feature_strains": dent_strains,
        "coincident_dent_feature": has_dent_coincident_feature,
        "note": (
            "B31.8 uses a 6% dent strain screening criterion. API RP 1183 requires shape, fatigue, coincident-feature, "
            "material, and uncertainty review; this workflow does not claim a DFDI calculation."
        ),
    }

    isolated_limits = []
    for item in screening["outputs"]["ranked_features"]:
        weld_result = weld_assessment_by_id.get(item["feature_id"])
        isolated_limits.append(
            {
                "source": item["feature_id"],
                "maximum_mop_mpa": weld_result["maximum_mop_mpa"] if weld_result else item["allowable_pressure_mpa"],
                "failure_pressure_mpa": weld_result["failure_pressure_mpa"] if weld_result else item["governing_failure_pressure_mpa"],
                "mode": weld_result["calculation_method"] if weld_result else item["calculation_method"],
            }
        )
    all_limits = isolated_limits + [
        {
            "source": item["model_id"],
            "maximum_mop_mpa": item["maximum_mop_mpa"],
            "failure_pressure_mpa": item["combined_failure_pressure_mpa"],
            "mode": item["governing_failure_mode"],
        }
        for item in pair_models
    ]
    governing_limit = min(all_limits, key=lambda item: item["maximum_mop_mpa"])
    crack_growth_lives = [
        item["fatigue_life_years"]
        for item in screening["outputs"]["ranked_features"]
        if item["fatigue_life_years"] is not None
    ] + [item["fatigue_life_years"] for item in pair_models if item["fatigue_life_years"] is not None]
    minimum_fatigue_life = min(crack_growth_lives) if crack_growth_lives else None
    maximum_mop = governing_limit["maximum_mop_mpa"]
    status = "ACTION REQUIRED" if maop > maximum_mop else "REVIEW REQUIRED" if maop > 0.8 * maximum_mop else "ACCEPTABLE"
    if strain_assessment["b31_8_status"] == "ACTION REQUIRED":
        status = "ACTION REQUIRED"
    refinement_multiplier = {"coarse": 0.7, "standard": 1.0, "fine": 1.5}.get(str(model.get("mesh_refinement", "standard")), 1.0)
    remote_element_size = max(diameter / (24.0 * refinement_multiplier), wall_thickness * 3.0)
    local_subdivision = {"coarse": 3.0, "standard": 4.0, "fine": 5.0}.get(str(model.get("mesh_refinement", "standard")), 4.0)
    local_element_size = remote_element_size / local_subdivision
    refinement_ratio = remote_element_size / local_element_size
    local_refinement_zones = [
        {
            "feature_id": feature["id"],
            "axial_center_mm": feature["axial_location_mm"],
            "clock_position_deg": feature["clock_position_deg"],
            "axial_extent_mm": max(feature["length_mm"] * 1.6, 8.0 * local_element_size),
            "circumferential_extent_mm": max(feature["width_mm"] * 1.6, 8.0 * local_element_size),
            "target_element_size_mm": local_element_size * (0.65 if feature["type"] == "crack" else 1.0),
            "refinement_reason": (
                "crack front and open faces"
                if feature["type"] == "crack"
                else "remaining ligament and local curvature"
                if feature["type"] == "metal_loss"
                else "dent curvature and shoulder strain"
                if feature["type"] == "dent"
                else "anomaly stress gradient"
            ),
        }
        for feature in translated
    ]
    raw_mesh_nodes = max(240, int((count * 260 + raw_sample_count * 72) * refinement_multiplier))
    crack_void_elements = sum(
        max(4, int((feature["effective_geometry"]["crack_front_points"] + 1) * 3 * refinement_multiplier))
        for feature in translated
        if feature["effective_geometry"]["mesh_void_required"]
    )
    raw_mesh_elements = max(360, int(raw_mesh_nodes * 1.85) - crack_void_elements)

    return {
        "inputs": {
            "feature_count": count,
            "assessment_factor": assessment_factor,
            "outside_diameter_mm": diameter,
            "wall_thickness_mm": wall_thickness,
            "interaction_distance_mm": interaction_distance,
            "mesh_refinement": str(model.get("mesh_refinement", "standard")),
            "geometry_source": geometry_source,
            "raw_sample_count": raw_sample_count,
            "surrogate_model": surrogate_name,
            "reliability_samples": reliability_samples,
            "reliability_enabled": reliability_enabled,
            "strain_limit": strain_limit,
            "weld_yield_strength_mpa": weld_yield,
            "weld_fracture_toughness_mpa_sqrt_m": weld_toughness,
            "weld_residual_stress_factor": weld_residual_factor,
            "weld_cap_width_mm": weld_cap_width,
            "haz_width_mm": haz_width,
            "pipe_seam_type": pipe_seam_type,
            "pipe_weld_type": pipe_weld_type,
            "manufacturing_process": manufacturing_process,
            "manufacturing_family": manufacturing_family,
            "spiral_turns": spiral_turns,
        },
        "translated_features": translated,
        "fea_models": pair_models,
        "screening": screening,
        "weld_assessment": {
            "feature_count": len(weld_assessments),
            "features": weld_assessments,
            "weld_strength_mismatch": weld_yield / smys,
            "pipe_seam_type": pipe_seam_type,
            "pipe_weld_type": pipe_weld_type,
            "manufacturing_process": manufacturing_process,
            "methodology": (
                "Weld location is resolved before capacity calculations. Crack-like weld and HAZ defects use a "
                "failure-assessment-diagram screen; other anomaly types use weld-zone capacity modifiers. "
                "Project-specific weld procedure, toughness, residual-stress, and flaw-characterization data are required "
                "for engineering acceptance."
            ),
        },
        "raw_mesh": {
            "mfl_samples": mfl_samples,
            "crack_samples": crack_samples,
            "caliper_samples": caliper_samples,
            "samples_by_feature": raw_by_feature,
            "node_count": raw_mesh_nodes,
            "element_count": raw_mesh_elements,
            "removed_crack_elements": crack_void_elements,
            "remote_element_size_mm": remote_element_size,
            "local_element_size_mm": local_element_size,
            "local_to_remote_density_ratio": refinement_ratio**2,
            "local_refinement_zones": local_refinement_zones,
            "adaptive_refinement": True,
            "generation_basis": (
                "Continuous shell topology with anomaly-centered cell subdivision; MFL-driven remaining-wall surface, "
                "caliper-driven radial shell deformation, and crack-front element removal"
            ),
            "geometry_coupled_to_assessment": True,
        },
        "surrogate": {
            "selected_model": ILI_SURROGATE_MODELS[surrogate_name]["label"],
            "predictions": surrogate_predictions,
            "validation": ili_surrogate_validation(surrogate_name),
            "governance": (
                "Advisory rapid prediction only. The lower confidence estimate is reported but does not replace "
                "the governing standards-based or validated nonlinear FEA result."
            ),
        },
        "reliability": reliability,
        "strain_assessment": strain_assessment,
        "workflow": [
            {"stage": 1, "name": "FEA automated reconstruction", "status": "Complete"},
            {"stage": 2, "name": "Weld and HAZ classification", "status": f"{len(weld_assessments)} weld-coincident feature(s)"},
            {"stage": 3, "name": "Adaptive FEA automation", "status": "Model ready"},
            {"stage": 4, "name": "ANN / DNN surrogate", "status": "Advisory prediction complete"},
            {
                "stage": 5,
                "name": "Reliability integration",
                "status": (
                    "Deactivated by user"
                    if not reliability_enabled
                    else "Probability assessment complete"
                    if reliability["available"]
                    else "Not applicable"
                ),
            },
            {"stage": 6, "name": "Strain acceptance", "status": f"B31.8: {b318_status}; API RP 1183: {api_1183_status}"},
        ],
        "outputs": {
            "status": status,
            "feature_count": count,
            "interaction_model_count": len(pair_models),
            "maximum_mop_mpa": maximum_mop,
            "current_maop_mpa": maop,
            "mop_utilization": maop / maximum_mop if maximum_mop > 0 else float("inf"),
            "governing_source": governing_limit["source"],
            "governing_failure_mode": governing_limit["mode"],
            "minimum_fatigue_life_years": minimum_fatigue_life,
            "critical_feature": screening["outputs"]["highest_risk_feature"],
            "raw_sample_count": raw_sample_count,
            "raw_mesh_nodes": raw_mesh_nodes,
            "raw_mesh_elements": raw_mesh_elements,
            "removed_crack_elements": crack_void_elements,
            "mesh_density_ratio": refinement_ratio**2,
            "surrogate_model": ILI_SURROGATE_MODELS[surrogate_name]["label"],
            "surrogate_minimum_mop_mpa": (
                min(item["advisory_maximum_mop_mpa"] for item in surrogate_predictions)
                if surrogate_predictions
                else None
            ),
            "reliability_enabled": reliability_enabled,
            "probability_of_failure": reliability["probability_of_failure"],
            "reliability_index_beta": reliability["reliability_index_beta"],
            "maximum_strain_percent": strain_assessment["maximum_strain_percent"],
            "b31_8_strain_status": b318_status,
            "api_rp_1183_status": api_1183_status,
            "weld_feature_count": len(weld_assessments),
            "disposition": (
                f"{count} ILI feature(s) translated into model geometry and {len(pair_models)} interaction model(s). "
                f"{len(weld_assessments)} feature(s) intersect weld metal or HAZ. "
                f"Maximum recommended MOP is {maximum_mop:.3f} MPa, governed by {governing_limit['source']}."
            ),
        },
    }


def calculate_prci_level2_dent_payload(payload: dict[str, Any]) -> dict[str, Any]:
    pipe = payload.get("pipe", {})
    dent = payload.get("dent", {})
    fatigue = payload.get("fatigue", {})
    crack_growth = payload.get("crack_growth", {})
    depth = positive_number(dent, "depth_mm")
    diameter = positive_number(pipe, "outside_diameter_mm")
    wall_thickness = positive_number(pipe, "wall_thickness_mm")
    pressure = positive_number(pipe, "operating_pressure_mpa")
    smys = positive_number(pipe, "smys_mpa")
    radius = positive_number(dent, "radius_mm")
    stress_cycles = positive_number(fatigue, "stress_cycles")
    scf = positive_number(fatigue, "stress_concentration_factor")

    dent_percent = depth / diameter * 100.0
    hoop_stress = pressure * diameter / (2.0 * wall_thickness)
    bending_strain = wall_thickness / (2.0 * radius)
    equivalent_stress = hoop_stress * scf
    if equivalent_stress <= 0:
        raise ValueError("equivalent stress must be greater than zero")
    fatigue_life = 1_000_000.0 / (equivalent_stress**1.5)
    remaining_strength_factor = smys / equivalent_stress

    criteria: list[dict[str, str]] = []

    def add_criterion(label: str, passes: bool, pass_message: str, fail_message: str) -> None:
        criteria.append(
            {
                "label": label,
                "status": "PASS" if passes else "FAIL",
                "message": pass_message if passes else fail_message,
            }
        )

    strain_limit = 0.06
    rsf_limit = 1.10
    add_criterion(
        "Bending strain",
        bending_strain <= strain_limit,
        f"Calculated bending strain {bending_strain:.4g} is at or below the {strain_limit:.4g} Level 2 strain limit.",
        f"Calculated bending strain {bending_strain:.4g} is greater than the {strain_limit:.4g} Level 2 strain limit; repair is required.",
    )
    add_criterion(
        "Remaining strength factor",
        remaining_strength_factor >= rsf_limit,
        f"Remaining strength factor {remaining_strength_factor:.3g} is at or above the {rsf_limit:.2f} screening limit.",
        f"Remaining strength factor {remaining_strength_factor:.3g} is below the {rsf_limit:.2f} screening limit; repair or further assessment is required.",
    )
    add_criterion(
        "Fatigue life",
        fatigue_life >= stress_cycles,
        f"Calculated fatigue life {fatigue_life:.0f} cycles is at or above the applied {stress_cycles:.0f} cycles.",
        f"Calculated fatigue life {fatigue_life:.0f} cycles is less than the applied {stress_cycles:.0f} cycles; repair or fatigue mitigation is required.",
    )

    result = "ACCEPTABLE"
    if bending_strain > strain_limit:
        result = "EXCESSIVE STRAIN - REPAIR REQUIRED"
    if remaining_strength_factor < rsf_limit:
        result = "INSUFFICIENT REMAINING STRENGTH"
    if fatigue_life < stress_cycles:
        result = "FATIGUE LIFE INSUFFICIENT"

    crack_growth_result: dict[str, Any] = {"enabled": False}
    if bool(crack_growth.get("enabled", False)):
        initial_crack = positive_number(crack_growth, "initial_crack_mm")
        final_crack = positive_number(crack_growth, "critical_crack_mm")
        delta_sigma = positive_number(crack_growth, "stress_range_mpa")
        paris_c = positive_number(crack_growth, "paris_c")
        paris_m = positive_number(crack_growth, "paris_m")
        increment = float(crack_growth.get("increment_mm", 0.001) or 0.001)
        if final_crack <= initial_crack:
            raise ValueError("critical_crack_mm must be greater than initial_crack_mm")
        if increment <= 0:
            raise ValueError("increment_mm must be greater than zero")
        current_crack = initial_crack
        cycles = 0.0
        steps = 0
        max_steps = 2_000_000
        while current_crack < final_crack and steps < max_steps:
            delta_k = delta_sigma * math.sqrt(math.pi * current_crack)
            crack_growth_rate = paris_c * (delta_k**paris_m)
            if crack_growth_rate <= 0:
                raise ValueError("Calculated crack growth rate must be greater than zero")
            step = min(increment, final_crack - current_crack)
            cycles += step / crack_growth_rate
            current_crack += step
            steps += 1
        if steps >= max_steps:
            raise ValueError("Fatigue crack growth integration did not converge. Increase increment size or check inputs.")
        crack_growth_result = {
            "enabled": True,
            "estimated_cycles": cycles,
            "final_crack_mm": final_crack,
            "initial_crack_mm": initial_crack,
            "steps": steps,
            "status": "PASS" if cycles >= stress_cycles else "FAIL",
        }
        add_criterion(
            "Crack growth life",
            cycles >= stress_cycles,
            f"Estimated crack growth life {cycles:.0f} cycles is at or above the applied {stress_cycles:.0f} cycles.",
            f"Estimated crack growth life {cycles:.0f} cycles is less than the applied {stress_cycles:.0f} cycles; repair or crack-growth mitigation is required.",
        )
        if cycles < stress_cycles:
            result = "FATIGUE CRACK GROWTH LIFE INSUFFICIENT"

    repair_reasons = [criterion["message"] for criterion in criteria if criterion["status"] == "FAIL"]
    acceptance_reasons = [criterion["message"] for criterion in criteria if criterion["status"] == "PASS"]

    return {
        "inputs": {
            "dent_depth_mm": depth,
            "outside_diameter_mm": diameter,
            "wall_thickness_mm": wall_thickness,
            "operating_pressure_mpa": pressure,
            "smys_mpa": smys,
            "dent_radius_mm": radius,
            "stress_cycles": stress_cycles,
            "stress_concentration_factor": scf,
        },
        "outputs": {
            "status": "ACCEPTABLE" if result == "ACCEPTABLE" else "REPAIR REQUIRED",
            "assessment_result": result,
            "dent_depth_percent": dent_percent,
            "hoop_stress_mpa": hoop_stress,
            "bending_strain": bending_strain,
            "equivalent_stress_mpa": equivalent_stress,
            "fatigue_life_cycles": fatigue_life,
            "remaining_strength_factor": remaining_strength_factor,
            "crack_growth": crack_growth_result,
            "criteria": criteria,
            "repair_required_reasons": repair_reasons,
            "acceptance_reasons": acceptance_reasons,
        },
    }
    effective_area = area_coefficients[crack_profile] * crack_depth * crack_length
    reference_area = crack_length * wall_thickness
    area_ratio = effective_area / reference_area
    failure_stress = flow_stress * ((1 - area_ratio) / (1 - effective_area / (folias_factor * reference_area)))
    collapse_pressure = 2 * wall_thickness * failure_stress / diameter

    qf = (
        1.2581
        - 0.20589 * (crack_depth / crack_length)
        - 11.493 * (crack_depth / crack_length) ** 2
        + 29.586 * (crack_depth / crack_length) ** 3
        - 23.584 * (crack_depth / crack_length) ** 4
    )
    fsf = (
        (2 * wall_thickness / (math.pi * crack_depth))
        * math.tan(math.pi * crack_depth / (2 * wall_thickness))
        * (1 - 2 * crack_depth / crack_length)
        + 2 * crack_depth / crack_length
    )
    strain_hardening = -0.00546 + 0.556 * (yield_strength / tensile_strength) - 0.547 * (yield_strength / tensile_strength) ** 2
    plastic_strain_yield = 0.005 - yield_strength / elastic_modulus
    if strain_hardening <= 0:
        raise ValueError("Calculated strain hardening exponent is not positive. Check yield and tensile strength.")
    if plastic_strain_yield <= 0:
        raise ValueError("Calculated plastic strain at yield is not positive. Check yield strength and elastic modulus.")

    pressure = 0.0
    iterations = 0
    hoop_stress = 0.0
    local_stress = 0.0
    plastic_strain = 0.0
    shih_hutchinson = 0.0
    elastic_j = 0.0
    plastic_j = 0.0
    total_j = 0.0
    stopped_by = "fracture"

    while True:
        hoop_stress = pressure * diameter / (2 * wall_thickness)
        crack_face_pressure = pressure * math.pi * crack_depth / (4 * wall_thickness) if crack_location == "internal" else 0.0
        local_stress = (hoop_stress + crack_face_pressure) * (
            (1 - math.pi * crack_depth / (4 * wall_thickness * folias_factor)) / (1 - math.pi * crack_depth / (4 * wall_thickness))
        )
        plastic_strain = plastic_strain_yield * (local_stress / yield_strength) ** (1 / strain_hardening)
        shih_hutchinson = ((3.85 * math.sqrt(1 / strain_hardening) * (1 - strain_hardening)) + math.pi * strain_hardening) * (
            1 + strain_hardening
        )
        elastic_j = qf * fsf * crack_depth * ((local_stress**2) * math.pi / elastic_modulus)
        plastic_j = qf * fsf * crack_depth * shih_hutchinson * plastic_strain * local_stress
        total_j = elastic_j + plastic_j
        if total_j >= fracture_toughness:
            stopped_by = "fracture"
            break
        pressure += pressure_step
        iterations += 1
        if pressure > collapse_pressure:
            stopped_by = "collapse"
            break
        if iterations >= max_iterations:
            stopped_by = "iteration_limit"
            break

    failure_pressure = min(pressure, collapse_pressure)
    controlling_mode = "Fracture" if pressure <= collapse_pressure else "Plastic collapse"
    psi_per_mpa = 145.037737796858
    return {
        "inputs": {
            "outside_diameter_mm": diameter,
            "wall_thickness_mm": wall_thickness,
            "crack_depth_mm": crack_depth,
            "crack_length_mm": crack_length,
            "yield_strength_mpa": yield_strength,
            "tensile_strength_mpa": tensile_strength,
            "elastic_modulus_mpa": elastic_modulus,
            "fracture_toughness_j": fracture_toughness,
            "fracture_toughness_method": toughness_method,
            "cvn_j": cvn,
            "charpy_area_in2": charpy_area,
            "crack_location": crack_location,
            "crack_profile": crack_profile,
            "flow_stress_coefficient": flow_coefficient,
            "pressure_step_mpa": pressure_step,
        },
        "outputs": {
            "fracture_pressure_mpa": pressure,
            "collapse_pressure_mpa": collapse_pressure,
            "failure_pressure_mpa": failure_pressure,
            "fracture_pressure_psi": pressure * psi_per_mpa,
            "collapse_pressure_psi": collapse_pressure * psi_per_mpa,
            "failure_pressure_psi": failure_pressure * psi_per_mpa,
            "controlling_mode": controlling_mode,
            "stopped_by": stopped_by,
            "iterations": iterations,
        },
        "intermediate": {
            "flow_stress_mpa": flow_stress,
            "folias_ratio": ratio,
            "folias_factor_m": folias_factor,
            "effective_flaw_area_mm2": effective_area,
            "effective_area_coefficient": area_coefficients[crack_profile],
            "reference_area_mm2": reference_area,
            "area_ratio": area_ratio,
            "failure_stress_mpa": failure_stress,
            "qf": qf,
            "fsf": fsf,
            "strain_hardening_n": strain_hardening,
            "plastic_strain_yield": plastic_strain_yield,
            "hoop_stress_mpa": hoop_stress,
            "local_stress_mpa": local_stress,
            "plastic_strain": plastic_strain,
            "shih_hutchinson_f3": shih_hutchinson,
            "elastic_j": elastic_j,
            "plastic_j": plastic_j,
            "total_j": total_j,
        },
    }


def fmt(value: Any, digits: int = 1) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if digits == 0:
        return f"{number:,.0f}"
    return f"{number:,.{digits}f}".rstrip("0").rstrip(".")


MITIGATION_MEDIA = {
    "none": {"stiffness": 0.0, "transferability": 1.0, "bypass": 0.0, "min_factor": 1.0, "description": "No mitigation"},
    "steel_plate": {
        "stiffness": 0.95,
        "transferability": 0.35,
        "bypass": 0.25,
        "min_factor": 0.03,
        "description": "stiff plate distribution with partial edge bypass",
    },
    "rig_mats": {
        "stiffness": 0.65,
        "transferability": 0.55,
        "bypass": 0.20,
        "min_factor": 0.08,
        "description": "moderate stiffness mat distribution",
    },
    "wooden_bridge": {
        "stiffness": 0.90,
        "transferability": 0.06,
        "bypass": 1.0,
        "min_factor": 0.0,
        "description": "bridge action transferring live load to supports away from the pipe",
    },
    "hollow_concrete_slab": {
        "stiffness": 0.85,
        "transferability": 0.22,
        "bypass": 0.50,
        "min_factor": 0.02,
        "description": "stiff slab distribution with partial structural bypass",
    },
    "custom": {
        "stiffness": 0.50,
        "transferability": 0.50,
        "bypass": 0.0,
        "min_factor": 0.05,
        "description": "custom load-spreading medium",
    },
}


def mitigation_result(payload: dict[str, Any], result: dict[str, Any]) -> dict[str, Any]:
    mitigation = payload.get("mitigation", {})
    mitigation_type = str(mitigation.get("type", "none"))
    width = float(mitigation.get("width_in", 0) or 0)
    length = float(mitigation.get("length_in", 0) or 0)
    thickness = float(mitigation.get("thickness_in", 0) or 0)
    spread_angle = max(0.0, min(float(mitigation.get("spread_angle_deg", 45) or 45), 89.0))
    unit_weight = float(mitigation.get("unit_weight_pcf", 0) or 0)
    if mitigation_type == "none" or width <= 0 or length <= 0:
        return {"applied": False, "factor": 1.0}
    media = MITIGATION_MEDIA.get(mitigation_type, MITIGATION_MEDIA["custom"])
    vehicle = payload.get("vehicle", {})
    contact_areas = [float(area) for area in vehicle.get("contact_areas_in2", [vehicle.get("contact_area_in2", 1)])]
    if str(vehicle.get("vehicle_type", "wheel")) == "track":
        loaded_contact_area = sum(max(area, 0.0) for area in contact_areas)
    else:
        loaded_contact_area = sum(max(area, 0.0) * 2.0 for area in contact_areas)
    spread = max(thickness, 0.0) * math.tan(math.radians(spread_angle))
    effective_area = max((width + 2.0 * spread) * (length + 2.0 * spread), 1.0)
    area_ratio = max(0.0, min(1.0, max(loaded_contact_area, 1.0) / effective_area))
    stiffness_area_factor = area_ratio ** max(float(media["stiffness"]), 0.05)
    pipe = payload.get("pipe", {})
    soil = payload.get("soil", {})
    pipe_influence_width = max(float(pipe.get("outside_diameter_in", 1) or 1) + 2.0 * float(soil.get("cover_in", 0) or 0), 1.0)
    pipe_influence_length = pipe_influence_width
    coverage_factor = max(0.0, min(1.0, width / pipe_influence_width) * min(1.0, length / pipe_influence_length))
    bypass_factor = max(0.0, min(1.0, float(media["bypass"]) * coverage_factor))
    factor = max(
        float(media["min_factor"]),
        min(1.0, stiffness_area_factor * float(media["transferability"]) * (1.0 - bypass_factor)),
    )
    self_weight_pressure = max(0.0, unit_weight) * max(0.0, thickness) / 1728.0
    live_pressure = max(float(result.get("critical_position", {}).get("live_pressure_psi", 0) or 0), 1e-9)
    self_weight_hoop = (self_weight_pressure / live_pressure) * result["mop"]["live_hoop_bending_psi"]
    live_hoop = result["mop"]["live_hoop_bending_psi"] * factor + self_weight_hoop
    hoop_stress = result["mop"]["total_hoop_stress_psi"] - result["mop"]["live_hoop_bending_psi"] + live_hoop
    assessment = result["mop"]["assessment_stress_psi"] - result["mop"]["live_hoop_bending_psi"] + live_hoop
    return {
        "applied": True,
        "factor": factor,
        "live_hoop_psi": live_hoop,
        "hoop_stress_psi": hoop_stress,
        "assessment_stress_psi": assessment,
        "effective_area_in2": effective_area,
        "effective_width_in": width + 2.0 * spread,
        "effective_length_in": length + 2.0 * spread,
        "loaded_contact_area_in2": loaded_contact_area,
        "area_ratio": area_ratio,
        "stiffness_area_factor": stiffness_area_factor,
        "bypass_factor": bypass_factor,
        "coverage_factor": coverage_factor,
        "self_weight_pressure_psi": self_weight_pressure,
        "self_weight_hoop_psi": self_weight_hoop,
        "media_description": str(media["description"]),
    }


def assessment_status(payload: dict[str, Any], result: dict[str, Any], mitigation: dict[str, Any]) -> dict[str, Any]:
    pipe = payload.get("pipe", {})
    pipe_meta = result["vehicle"]["pipe"]
    smys = max(float(pipe.get("smys_psi", 0) or 0), 1e-9)
    mitigation_applied = bool(mitigation.get("applied"))
    assessment_stress = (
        float(mitigation.get("assessment_stress_psi", result["mop"]["assessment_stress_psi"]))
        if mitigation_applied
        else float(result["mop"]["assessment_stress_psi"])
    )
    assessment_percent = 100.0 * assessment_stress / smys
    passes = assessment_percent <= float(pipe_meta["hoop_limit_percent_smys"])
    bending_strain = abs(float(result["mop"].get("pre_existing_bending_strain_microstrain", 0) or 0))
    strain_phrase = " including pre-existing bending strain" if bending_strain > 1e-6 else ""
    limit_phrase = (
        f"CSA Z662 Class {pipe_meta['class_location']} limit: "
        f"{fmt(pipe_meta['hoop_limit_percent_smys'], 1)}% SMYS ({fmt(pipe_meta['hoop_limit_psi'])} psi)."
    )
    if mitigation_applied:
        mitigation_label = str(payload.get("mitigation", {}).get("type", "mitigation")).replace("_", " ")
        detail = (
            f"{fmt(assessment_percent, 2)}% of SMYS at MAOP after {mitigation_label} mitigation{strain_phrase}. "
            f"Unmitigated: {fmt(result['mop']['assessment_percent_smys'], 2)}% of SMYS. {limit_phrase}"
        )
    else:
        detail = f"{fmt(result['mop']['assessment_percent_smys'], 2)}% of SMYS at MAOP{strain_phrase}. {limit_phrase}"
    return {
        "passes": passes,
        "status_text": "PASS" if passes else "FAIL",
        "basis_label": "Assessment stress",
        "assessment_stress_psi": assessment_stress,
        "assessment_percent_smys": assessment_percent,
        "detail": detail,
        "mitigation_applied": mitigation_applied,
    }


def fatigue_result(payload: dict[str, Any], result: dict[str, Any]) -> dict[str, Any]:
    fatigue = payload.get("fatigue", {})
    if not fatigue.get("enabled"):
        return {"status": "Not checked", "damage_ratio": 0.0, "allowable_cycles": 0.0, "stress_range_psi": 0.0}
    stress_range = (
        float(fatigue.get("manual_stress_range_psi", 0) or 0)
        if fatigue.get("stress_source") == "manual"
        else float(result["mop"]["live_hoop_bending_psi"])
    )
    cycles = float(fatigue.get("applied_cycles", 0) or 0)
    exponent = float(fatigue.get("exponent", 3) or 3)
    constant = float(fatigue.get("constant", 0) or 0)
    if stress_range <= 0 or cycles <= 0 or exponent <= 0 or constant <= 0:
        return {"status": "Incomplete", "damage_ratio": 0.0, "allowable_cycles": 0.0, "stress_range_psi": stress_range}
    allowable = constant / stress_range**exponent
    damage = cycles / allowable if allowable > 0 else float("inf")
    return {
        "status": "PASS" if damage <= 1.0 else "FAIL",
        "damage_ratio": damage,
        "allowable_cycles": allowable,
        "stress_range_psi": stress_range,
    }


def table_story(title: str, rows: list[tuple[str, Any]], styles: dict[str, Any]) -> list[Any]:
    story: list[Any] = [Paragraph(title, styles["Heading2"])]
    table = Table([[Paragraph(str(label), styles["BodyText"]), Paragraph(str(value), styles["BodyText"])] for label, value in rows], colWidths=[2.35 * inch, 4.35 * inch])
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d8dee4")),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f4f6f8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend([table, Spacer(1, 0.14 * inch)])
    return story


def build_report_pdf(payload: dict[str, Any]) -> bytes:
    result = calculate_gui_payload(payload)
    mitigation = mitigation_result(payload, result)
    fatigue = fatigue_result(payload, result)
    pipe = payload["pipe"]
    vehicle = payload["vehicle"]
    soil = payload["soil"]
    strain = payload.get("strain", {})
    pipe_meta = result["vehicle"]["pipe"]
    status = assessment_status(payload, result, mitigation)
    passes = status["passes"]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph("Pipeline Crossing Assessment Report", styles["Title"]),
        Paragraph(f"Assessment: {payload.get('assessment_name', 'Untitled assessment')}", styles["Normal"]),
        Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]),
        Spacer(1, 0.18 * inch),
    ]
    status_color = colors.HexColor("#14532d" if passes else "#7f1d1d")
    status_table = Table(
        [[Paragraph("ASSESSMENT RESULT", styles["BodyText"]), Paragraph(str(status["status_text"]), styles["Title"])]],
        colWidths=[2.0 * inch, 4.7 * inch],
    )
    status_table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 2, colors.HexColor("#16a34a" if passes else "#dc2626")),
                ("TEXTCOLOR", (1, 0), (1, 0), status_color),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("PADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.extend([status_table, Paragraph(str(status["detail"]), styles["Normal"]), Spacer(1, 0.14 * inch)])

    calculated_rows = [
        (f"{status['basis_label']} used for PASS/FAIL", f"{fmt(status['assessment_stress_psi'])} psi"),
        (f"{status['basis_label']} ratio used for PASS/FAIL", f"{fmt(status['assessment_percent_smys'], 2)}% of SMYS"),
        ("Mitigation included in PASS/FAIL", str(payload.get("mitigation", {}).get("type", "none")).replace("_", " ") if mitigation.get("applied") else "No"),
        ("Unmitigated hoop stress at MAOP", f"{fmt(result['mop']['total_hoop_stress_psi'])} psi"),
        ("Unmitigated hoop stress ratio at MAOP", f"{fmt(result['mop']['total_hoop_percent_smys'], 2)}% of SMYS"),
        ("Assessment stress including bending strain", f"{fmt(result['mop']['assessment_stress_psi'])} psi"),
        ("Assessment stress ratio including bending strain", f"{fmt(result['mop']['assessment_percent_smys'], 2)}% of SMYS"),
        ("CSA Z662 class limit", f"{fmt(pipe_meta['hoop_limit_percent_smys'], 1)}% SMYS ({fmt(pipe_meta['hoop_limit_psi'])} psi)"),
        ("Live load hoop", f"{fmt(result['mop']['live_hoop_bending_psi'])} psi"),
        ("Hoop stress subtotal", f"{fmt(result['mop']['total_hoop_stress_psi'])} psi"),
        ("Bending strain stress", f"{fmt(result['mop']['pre_existing_bending_stress_psi'])} psi"),
        ("Soil hoop", f"{fmt(result['mop']['soil_hoop_bending_psi'])} psi"),
        ("Pressure hoop", f"{fmt(result['mop']['pressure_hoop_psi'])} psi"),
        ("Fatigue status", fatigue["status"]),
        ("Fatigue damage ratio", fmt(fatigue["damage_ratio"], 4)),
        ("Mitigated hoop stress at MAOP", f"{fmt(mitigation.get('hoop_stress_psi'))} psi" if mitigation.get("applied") else "Not applied"),
        ("Mitigated assessment including bending strain", f"{fmt(mitigation.get('assessment_stress_psi'))} psi" if mitigation.get("applied") else "Not applied"),
        ("Mitigation load factor", fmt(mitigation.get("factor", 1), 4)),
        ("Mitigation effective area", f"{fmt(mitigation.get('effective_area_in2', 0), 1)} in2" if mitigation.get("applied") else "Not applied"),
        ("Mitigation bypass", f"{fmt(100 * mitigation.get('bypass_factor', 0), 1)}%" if mitigation.get("applied") else "Not applied"),
        ("Mitigation self-weight hoop", f"{fmt(mitigation.get('self_weight_hoop_psi', 0), 1)} psi" if mitigation.get("applied") else "Not applied"),
    ]
    story.extend(table_story("Calculated Outputs", calculated_rows, styles))

    story.extend(
        table_story(
            "Pipeline Inputs",
            [
                ("Outside diameter", f"{fmt(pipe['outside_diameter_in'], 3)} in"),
                ("Wall thickness", f"{fmt(pipe['wall_thickness_in'], 3)} in"),
                ("MAOP", f"{fmt(pipe['maop_psig'], 0)} psig"),
                ("Class location", pipe.get("class_location", "")),
                ("Design factor", pipe.get("design_factor", "")),
                ("SMYS", f"{fmt(pipe['smys_psi'], 0)} psi"),
                ("Young's modulus", f"{fmt(pipe['youngs_modulus_psi'], 0)} psi"),
            ],
            styles,
        )
    )
    story.extend(
        table_story(
            "Vehicle Inputs",
            [
                ("Vehicle type", vehicle.get("vehicle_type", "")),
                ("Number of axles/stations", vehicle.get("axle_count", "")),
                ("Vehicle weight", f"{fmt(sum(float(x) for x in vehicle.get('axle_loads_lb', [])), 0)} lb"),
                ("Contact area", f"{fmt(vehicle.get('contact_area_in2', 0), 1)} in2"),
                ("Axle width", f"{fmt(vehicle.get('axle_width_in', 0), 1)} in"),
                ("Crossing angle", f"{fmt(vehicle.get('crossing_angle_deg', 0), 1)} deg"),
                ("Impact factor", vehicle.get("impact_factor", "")),
            ],
            styles,
        )
    )
    story.extend(
        table_story(
            "Soil, Strain, And Mitigation Inputs",
            [
                ("Soil profile", soil.get("profile", "")),
                ("Soil load equation", soil.get("load_model", "")),
                ("Depth of cover", f"{fmt(soil.get('cover_in', 0), 1)} in"),
                ("Bending strain", f"{fmt(strain.get('bending_strain_microstrain', 0), 1)} microstrain"),
                ("Mitigation type", payload.get("mitigation", {}).get("type", "none")),
                ("Mitigation width", f"{fmt(payload.get('mitigation', {}).get('width_in', 0), 1)} in"),
                ("Mitigation length", f"{fmt(payload.get('mitigation', {}).get('length_in', 0), 1)} in"),
                ("Mitigation thickness", f"{fmt(payload.get('mitigation', {}).get('thickness_in', 0), 2)} in"),
            ],
            styles,
        )
    )
    story.append(Paragraph("Developed by M. Razzak. All rights reserved.", styles["Normal"]))
    doc.build(story)
    return buffer.getvalue()


class GuiRequestHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        path = self.path.split("?", 1)[0]
        if path == "/api/health":
            self.send_json(200, {"ok": True, "status": "healthy"})
            return
        if path == "/api/runtime-config":
            no_login = str(os.environ.get("PIPELINE_ASSESSMENT_NO_LOGIN", "")).strip().lower() in {"1", "true", "yes", "on"}
            self.send_json(
                200,
                {
                    "ok": True,
                    "no_login": no_login,
                    "local_user": os.environ.get("PIPELINE_ASSESSMENT_LOCAL_USER", "local-user"),
                },
            )
            return
        if path.startswith("/api/report-download/"):
            filename = path.rsplit("/", 1)[-1]
            file_path = (REPORT_ROOT / filename).resolve()
            if not str(file_path).startswith(str(REPORT_ROOT.resolve())) or not file_path.is_file():
                self.send_error(404)
                return
            content = file_path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
            return
        if path == "/":
            path = "/index.html"
        file_path = (WEB_ROOT / path.lstrip("/")).resolve()
        if not str(file_path).startswith(str(WEB_ROOT.resolve())) or not file_path.is_file():
            self.send_error(404)
            return
        content = file_path.read_bytes()
        content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def do_POST(self) -> None:
        try:
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            parsed_url = urlparse(self.path)
            request_path = parsed_url.path
            if request_path == "/api/ili-feature-import":
                filename, file_content = parse_multipart_file(self.headers.get("Content-Type", ""), body)
                features = parse_ili_feature_file(filename, file_content)
                self.send_json(200, {"ok": True, "filename": filename, "features": features, "count": len(features["ids"])})
                return
            if request_path == "/api/ili-raw-import":
                tool = parse_qs(parsed_url.query).get("tool", [""])[0]
                filename, file_content = parse_multipart_file(self.headers.get("Content-Type", ""), body)
                samples = parse_ili_raw_file(filename, file_content, tool)
                self.send_json(200, {"ok": True, "filename": filename, "tool": tool, "samples": samples, "count": len(samples)})
                return
            payload = json.loads(body.decode("utf-8"))
            if request_path == "/api/login":
                users = read_users()
                username = str(payload.get("username", "")).strip()
                password = str(payload.get("password", ""))
                user = users.get(username)
                if not user or not verify_password(password, str(user.get("password_hash", ""))):
                    self.send_json(401, {"ok": False, "error": "Invalid username or password."})
                    return
                if account_is_expired(user):
                    self.send_json(403, {"ok": False, "error": "This user login has expired. Contact an administrator."})
                    return
                if password_is_expired(user):
                    self.send_json(
                        403,
                        {
                            "ok": False,
                            "error": "Password expired. Change your password to continue.",
                            "require_password_change": True,
                            "username": username,
                        },
                    )
                    return
                session_id = start_user_session(users, username)
                write_users(users)
                self.send_json(200, {"ok": True, "user": public_user(username, users[username]), "session_id": session_id})
                return
            if request_path == "/api/session-restore":
                self.handle_session_restore(payload)
                return
            if request_path == "/api/session-activity":
                self.handle_session_activity(payload)
                return
            if request_path == "/api/change-password":
                self.handle_change_password(payload)
                return
            if request_path == "/api/admin/config":
                self.handle_admin_config(payload)
                return
            if request_path == "/api/admin/user":
                self.handle_admin_user(payload)
                return
            if request_path == "/api/admin/delete-user":
                self.handle_admin_delete_user(payload)
                return
            if request_path == "/api/admin/method":
                self.handle_admin_method(payload)
                return
            if request_path == "/api/report":
                pdf = build_report_pdf(payload)
                filename = str(payload.get("assessment_name", "pipeline-crossing-assessment")).strip() or "pipeline-crossing-assessment"
                safe = "".join(ch if ch.isalnum() else "-" for ch in filename).strip("-").lower() or "pipeline-crossing-assessment"
                self.send_response(200)
                self.send_header("Content-Type", "application/pdf")
                self.send_header("Content-Disposition", f'attachment; filename="{safe}-report.pdf"')
                self.send_header("Content-Length", str(len(pdf)))
                self.end_headers()
                self.wfile.write(pdf)
                return
            if request_path == "/api/report-link":
                pdf = build_report_pdf(payload)
                filename = str(payload.get("assessment_name", "pipeline-crossing-assessment")).strip() or "pipeline-crossing-assessment"
                safe = "".join(ch if ch.isalnum() else "-" for ch in filename).strip("-").lower() or "pipeline-crossing-assessment"
                report_name = f"{safe}-{secrets.token_hex(6)}-report.pdf"
                report_path = REPORT_ROOT / report_name
                report_path.write_bytes(pdf)
                self.send_json(200, {"ok": True, "url": f"/api/report-download/{report_name}", "filename": f"{safe}-report.pdf", "bytes": len(pdf)})
                return
            if request_path == "/api/corlas-calculate":
                self.send_json(200, {"ok": True, "result": calculate_corlas_payload(payload)})
                return
            if request_path == "/api/annex-k-eca-calculate":
                self.send_json(200, {"ok": True, "result": calculate_annex_k_eca_payload(payload)})
                return
            if request_path == "/api/dent-assessment-calculate":
                self.send_json(200, {"ok": True, "result": calculate_dent_assessment_payload(payload)})
                return
            if request_path == "/api/modified-b31g-calculate":
                self.send_json(200, {"ok": True, "result": calculate_modified_b31g_payload(payload)})
                return
            if request_path == "/api/rstreng-calculate":
                self.send_json(200, {"ok": True, "result": calculate_rstreng_payload(payload)})
                return
            if request_path == "/api/scc-colony-calculate":
                self.send_json(200, {"ok": True, "result": calculate_scc_colony_payload(payload)})
                return
            if request_path == "/api/crack-growth-calculate":
                self.send_json(200, {"ok": True, "result": calculate_crack_growth_payload(payload)})
                return
            if request_path == "/api/ili-screening-calculate":
                self.send_json(200, {"ok": True, "result": calculate_ili_screening_payload(payload)})
                return
            if request_path == "/api/prci-level2-dent-calculate":
                self.send_json(200, {"ok": True, "result": calculate_prci_level2_dent_payload(payload)})
                return
            if request_path == "/api/interacting-anomalies-calculate":
                self.send_json(200, {"ok": True, "result": calculate_interacting_anomalies_payload(payload)})
                return
            if request_path == "/api/ili-to-fea-calculate":
                self.send_json(200, {"ok": True, "result": calculate_ili_to_fea_payload(payload)})
                return
            if request_path != "/api/calculate":
                self.send_error(404)
                return
            response = {"ok": True, "result": calculate_gui_payload(payload)}
            self.send_json(200, response)
        except Exception as exc:
            self.send_json(400, {"ok": False, "error": str(exc)})

    def require_admin(self, payload: dict[str, Any]) -> None:
        users = read_users()
        username = str(payload.get("admin_username", "")).strip()
        password = str(payload.get("admin_password", ""))
        session_id = str(payload.get("session_id", "")).strip()
        user = users.get(username)
        if password:
            if not user or not verify_password(password, str(user.get("password_hash", ""))) or user.get("role") != "admin":
                raise ValueError("Admin username and password are required.")
        elif not user or user.get("role") != "admin" or not session_is_active(user, session_id):
            raise ValueError("An active admin login session is required.")
        if account_is_expired(user):
            raise ValueError("Admin account has expired.")

    def handle_session_activity(self, payload: dict[str, Any]) -> None:
        users = read_users()
        username = str(payload.get("username", "")).strip()
        session_id = str(payload.get("session_id", "")).strip()
        module = str(payload.get("module", "")).strip()
        elapsed_seconds = float(payload.get("elapsed_seconds", 0.0) or 0.0)
        finished = bool(payload.get("finished", False))
        if not username or not session_id:
            raise ValueError("username and session_id are required")
        if username not in users or not session_is_active(users[username], session_id):
            raise ValueError("An active login session is required.")
        user = record_session_activity(users, username, session_id, elapsed_seconds, module, finished)
        write_users(users)
        self.send_json(200, {"ok": True, "user": public_user(username, user)})

    def handle_session_restore(self, payload: dict[str, Any]) -> None:
        users = read_users()
        username = str(payload.get("username", "")).strip()
        session_id = str(payload.get("session_id", "")).strip()
        if not username or not session_id:
            raise ValueError("username and session_id are required")
        user = users.get(username)
        if not user or not session_is_active(user, session_id):
            self.send_json(401, {"ok": False, "error": "Stored login session is no longer active."})
            return
        if account_is_expired(user):
            self.send_json(403, {"ok": False, "error": "This user login has expired. Contact an administrator."})
            return
        if password_is_expired(user):
            self.send_json(403, {"ok": False, "error": "Password expired. Login again to change your password."})
            return
        user = record_session_activity(users, username, session_id, 0.0)
        write_users(users)
        self.send_json(200, {"ok": True, "user": public_user(username, user), "session_id": session_id})

    def handle_change_password(self, payload: dict[str, Any]) -> None:
        users = read_users()
        username = str(payload.get("username", "")).strip()
        current_password = str(payload.get("current_password", ""))
        new_password = str(payload.get("new_password", ""))
        user = users.get(username)
        if not user or not verify_password(current_password, str(user.get("password_hash", ""))):
            raise ValueError("Current username or password is incorrect.")
        if account_is_expired(user):
            raise ValueError("This user login has expired. Contact an administrator.")
        validate_password_strength(new_password)
        if verify_password(new_password, str(user.get("password_hash", ""))):
            raise ValueError("New password must be different from the current password.")
        user["password_hash"] = hash_password(new_password)
        user["password_changed_at"] = iso_date(today_utc())
        user["password_expires_at"] = password_expiry_from()
        users[username] = user
        write_users(users)
        self.send_json(200, {"ok": True, "user": public_user(username, user)})

    def handle_admin_user(self, payload: dict[str, Any]) -> None:
        self.require_admin(payload)
        users = read_users()
        username = str(payload.get("username", "")).strip()
        password = str(payload.get("password", ""))
        role = str(payload.get("role", "user")).strip() or "user"
        full_name = str(payload.get("full_name", "")).strip()
        email = str(payload.get("email", "")).strip()
        account_expires_at = str(payload.get("account_expires_at", "")).strip()
        if account_expires_at:
            parse_iso_date(account_expires_at)
        if not username:
            raise ValueError("username is required")
        validate_email(email)
        if role not in {"user", "admin"}:
            raise ValueError("role must be user or admin")
        existing = users.get(username)
        if existing:
            user = dict(existing)
            if password:
                validate_password_strength(password)
                user["password_hash"] = hash_password(password)
                user["password_changed_at"] = iso_date(today_utc())
                user["password_expires_at"] = password_expiry_from()
            user["role"] = role
            user["full_name"] = full_name
            user["email"] = email
            user["account_expires_at"] = account_expires_at
        else:
            if not password:
                raise ValueError("password is required for a new user")
            user = make_user_record(password, role, account_expires_at, full_name, email)
        users[username] = user
        write_users(users)
        self.send_json(200, {"ok": True, "users": public_users(users)})

    def handle_admin_delete_user(self, payload: dict[str, Any]) -> None:
        self.require_admin(payload)
        users = read_users()
        username = str(payload.get("username", "")).strip()
        admin_username = str(payload.get("admin_username", "")).strip()
        if not username:
            raise ValueError("username is required")
        if username == admin_username:
            raise ValueError("Administrators cannot delete their own account while signed in.")
        if username not in users:
            raise ValueError("User was not found.")
        del users[username]
        write_users(users)
        self.send_json(200, {"ok": True, "users": public_users(users)})

    def handle_admin_config(self, payload: dict[str, Any]) -> None:
        self.require_admin(payload)
        users = read_users()
        self.send_json(200, {"ok": True, "config": read_method_config(), "users": public_users(users)})

    def handle_admin_method(self, payload: dict[str, Any]) -> None:
        self.require_admin(payload)
        config = read_method_config()
        method = payload.get("method", {})
        for key in ("method_name", "notes"):
            if key in method:
                config[key] = str(method[key])
        for key in ("fatigue_constant", "fatigue_exponent"):
            if key in method:
                value = float(method[key])
                if value <= 0:
                    raise ValueError(f"{key} must be greater than zero")
                config[key] = value
        write_method_config(config)
        self.send_json(200, {"ok": True, "config": config})

    def send_json(self, status: int, payload: dict[str, Any]) -> None:
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def end_headers(self) -> None:
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "SAMEORIGIN")
        self.send_header("Referrer-Policy", "same-origin")
        self.send_header("Permissions-Policy", "geolocation=(), microphone=(), camera=()")
        self.send_header(
            "Content-Security-Policy",
            "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; frame-src 'self'; object-src 'none'; base-uri 'self'",
        )
        super().end_headers()

    def log_message(self, format: str, *args: Any) -> None:
        return


def run(host: str | None = None, port: int | None = None) -> None:
    ensure_data_files()
    host = host or os.environ.get("PIPELINE_ASSESSMENT_HOST", "127.0.0.1")
    port = port or int(os.environ.get("PIPELINE_ASSESSMENT_PORT") or os.environ.get("PORT") or "8765")
    server = ThreadingHTTPServer((host, port), GuiRequestHandler)
    print(f"CEPA Crossing GUI running at http://{host}:{port}/")
    server.serve_forever()


def main() -> None:
    run()


if __name__ == "__main__":
    main()
